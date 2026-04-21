from __future__ import annotations

import json
import math
import os
import random
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Set, Tuple

os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
LAB_GROUP_COUNT = 3
TimetableCell = Dict[str, object]
TimetableMatrix = List[List[TimetableCell | None]]
SectionTable = Dict[str, object]
ClassTable = Dict[str, object]
FacultyTable = Dict[str, object]
RoomTable = Dict[str, object]
FacultyAssignment = Dict[str, object]


def create_time_slots(working_days: int, periods_per_day: int) -> List[Tuple[int, int]]:
    """Create slot coordinates as (day_index, period_index)."""
    return [(day, period) for day in range(working_days) for period in range(periods_per_day)]


def build_section_key(class_name: str, standard: str, section: str) -> str:
    return f"{class_name} | Std {standard} | Sec {section}"


def build_class_key(class_name: str, standard: str) -> str:
    return f"{class_name} | Std {standard}"


def _extract_section_group(section: str) -> Tuple[str, int | None]:
    cleaned = str(section).strip()
    upper_cleaned = cleaned.upper()
    marker = "-G"
    marker_index = upper_cleaned.rfind(marker)
    if marker_index <= 0:
        return cleaned, None

    suffix = upper_cleaned[marker_index + len(marker):]
    if suffix in {str(index) for index in range(1, LAB_GROUP_COUNT + 1)}:
        return cleaned[:marker_index], int(suffix)

    return cleaned, None


def _build_group_section_code(base_section: str, group_number: int) -> str:
    return f"{base_section}-G{group_number}"


def _build_section_family_key(class_name: str, standard: str, section: str) -> str:
    base_section, _ = _extract_section_group(section)
    return build_section_key(class_name, standard, base_section)


def _can_assign_distinct_rooms(room_options: List[List[str]]) -> bool:
    if not room_options:
        return True

    normalized_options: List[List[str]] = []
    for raw_options in room_options:
        unique_options: List[str] = []
        seen: Set[str] = set()
        for room_name in raw_options:
            normalized_room = str(room_name).strip()
            if not normalized_room or normalized_room in seen:
                continue
            seen.add(normalized_room)
            unique_options.append(normalized_room)

        if not unique_options:
            return False
        normalized_options.append(unique_options)

    member_order = sorted(
        range(len(normalized_options)),
        key=lambda member_index: len(normalized_options[member_index]),
    )
    used_rooms: Set[str] = set()

    def backtrack(order_index: int) -> bool:
        if order_index >= len(member_order):
            return True

        member_index = member_order[order_index]
        for room_name in normalized_options[member_index]:
            if room_name in used_rooms:
                continue

            used_rooms.add(room_name)
            if backtrack(order_index + 1):
                return True
            used_rooms.remove(room_name)

        return False

    return backtrack(0)


def _pick_lab_triplet_with_distinct_faculty(
    session_units: List[Dict[str, object]],
) -> List[Dict[str, object]] | None:
    if len(session_units) < 3:
        return None

    best_triplet: List[Dict[str, object]] | None = None
    best_score: Tuple[int, int, int, int] | None = None

    for first_index in range(len(session_units) - 2):
        first = session_units[first_index]
        first_teacher = str(first["teacher"])

        for second_index in range(first_index + 1, len(session_units) - 1):
            second = session_units[second_index]
            second_teacher = str(second["teacher"])
            if second_teacher == first_teacher:
                continue

            for third_index in range(second_index + 1, len(session_units)):
                third = session_units[third_index]
                third_teacher = str(third["teacher"])
                if third_teacher in {first_teacher, second_teacher}:
                    continue

                triplet = [first, second, third]
                room_options = [list(item.get("allowed_rooms", [])) for item in triplet]
                if not _can_assign_distinct_rooms(room_options):
                    continue

                subject_diversity = len({str(item["subject"]) for item in triplet})
                room_option_count = sum(len(item.get("allowed_rooms", [])) for item in triplet)
                distinct_room_variety = len(
                    {
                        str(room_name).strip()
                        for options in room_options
                        for room_name in options
                        if str(room_name).strip()
                    }
                )
                source_order_score = -sum(int(item.get("source_index", 0)) for item in triplet)
                score = (
                    subject_diversity,
                    distinct_room_variety,
                    room_option_count,
                    source_order_score,
                )

                if best_score is None or score > best_score:
                    best_score = score
                    best_triplet = triplet

    return best_triplet


def _expand_assignments_for_lab_groups(
    assignments: List[FacultyAssignment],
) -> List[FacultyAssignment]:
    expanded_assignments: List[FacultyAssignment] = []
    lab_sessions_by_base_section: Dict[Tuple[str, str, str], List[Dict[str, object]]] = {}

    for source_index, assignment in enumerate(assignments):
        class_name = str(assignment["class_name"])
        standard = str(assignment["standard"])
        raw_section = str(assignment["section"])
        section, explicit_group = _extract_section_group(raw_section)
        subject_name = str(assignment["subject"])
        teacher_name = str(assignment["teacher"])
        lectures = int(assignment["lectures"])
        is_lab = bool(assignment.get("is_lab", False))
        allowed_rooms = [str(room) for room in assignment.get("allowed_rooms", []) if str(room).strip()]
        subject_daily_max = assignment.get("subject_daily_max")
        lab_bundle_order_raw = assignment.get("lab_bundle_order")
        lab_subject_count_raw = assignment.get("lab_subject_count")

        normalized_assignment: FacultyAssignment = {
            "class_name": class_name,
            "standard": standard,
            "section": raw_section,
            "subject": subject_name,
            "teacher": teacher_name,
            "lectures": lectures,
            "is_lab": is_lab,
            "allowed_rooms": allowed_rooms,
            "subject_daily_max": subject_daily_max,
        }

        if not is_lab:
            expanded_assignments.append(normalized_assignment)
            continue

        if lab_bundle_order_raw is not None:
            try:
                normalized_assignment["lab_bundle_order"] = int(lab_bundle_order_raw)
            except (TypeError, ValueError):
                pass

        if lab_subject_count_raw is not None:
            try:
                normalized_assignment["lab_subject_count"] = int(lab_subject_count_raw)
            except (TypeError, ValueError):
                pass

        if not allowed_rooms:
            raise ValueError(
                f"LAB assignment {class_name}-{standard}-{section} -> {subject_name} has no allowed rooms."
            )

        if explicit_group is not None:
            bundle_key = str(assignment.get("lab_bundle_key", "")).strip()
            if bundle_key:
                normalized_assignment["lab_bundle_key"] = bundle_key
            expanded_assignments.append(normalized_assignment)
            continue

        if lectures % 2 != 0:
            raise ValueError(
                f"LAB assignment {class_name}-{standard}-{section} -> {subject_name} has odd lectures ({lectures})."
            )

        session_count = lectures // 2
        if session_count < 1:
            raise ValueError(
                f"LAB assignment {class_name}-{standard}-{section} -> {subject_name} must have at least 2 lectures."
            )

        # Interpret base-section LAB load as per group and replicate it to all groups.
        session_count *= LAB_GROUP_COUNT

        section_key = (class_name, standard, section)
        section_sessions = lab_sessions_by_base_section.setdefault(section_key, [])
        for session_index in range(session_count):
            section_sessions.append(
                {
                    "class_name": class_name,
                    "standard": standard,
                    "base_section": section,
                    "subject": subject_name,
                    "teacher": teacher_name,
                    "allowed_rooms": allowed_rooms,
                    "subject_daily_max": subject_daily_max,
                    "source_index": source_index,
                    "session_index": session_index,
                }
            )

    for class_name, standard, base_section in sorted(lab_sessions_by_base_section.keys()):
        remaining_sessions = list(lab_sessions_by_base_section[(class_name, standard, base_section)])
        section_subject_count = len({str(item.get("subject", "")).strip() for item in remaining_sessions})

        if len(remaining_sessions) % LAB_GROUP_COUNT != 0:
            raise ValueError(
                f"{build_section_key(class_name, standard, base_section)} has {len(remaining_sessions)} LAB sessions. "
                f"LAB group splitting requires multiples of {LAB_GROUP_COUNT} sessions."
            )

        bundle_index = 1
        while remaining_sessions:
            triplet = _pick_lab_triplet_with_distinct_faculty(remaining_sessions)
            if triplet is None:
                raise ValueError(
                    f"{build_section_key(class_name, standard, base_section)} LAB sessions cannot be split into "
                    f"triplets with {LAB_GROUP_COUNT} distinct faculty and {LAB_GROUP_COUNT} distinct allowed rooms."
                )

            for session in triplet:
                remaining_sessions.remove(session)

            triplet.sort(
                key=lambda item: (
                    int(item.get("source_index", 0)),
                    int(item.get("session_index", 0)),
                    str(item.get("subject", "")),
                    str(item.get("teacher", "")),
                )
            )

            # Rotate subject-to-group mapping by bundle index so groups alternate subjects over time.
            rotation_offset = (bundle_index - 1) % LAB_GROUP_COUNT
            rotated_triplet = triplet[rotation_offset:] + triplet[:rotation_offset]

            bundle_key = (
                f"{build_section_key(class_name, standard, base_section)}::LAB-BUNDLE-{bundle_index}"
            )
            for group_number, session in enumerate(rotated_triplet, start=1):
                expanded_assignments.append(
                    {
                        "class_name": class_name,
                        "standard": standard,
                        "section": _build_group_section_code(base_section, group_number),
                        "subject": str(session["subject"]),
                        "teacher": str(session["teacher"]),
                        "lectures": 2,
                        "is_lab": True,
                        "allowed_rooms": list(session.get("allowed_rooms", [])),
                        "subject_daily_max": session.get("subject_daily_max"),
                        "lab_bundle_key": bundle_key,
                        "lab_bundle_order": bundle_index,
                        "lab_subject_count": section_subject_count,
                    }
                )

            bundle_index += 1

    return expanded_assignments


def _cells_form_single_lab_block(first_cell: TimetableCell, second_cell: TimetableCell) -> bool:
    if not bool(first_cell.get("is_lab", False)) or not bool(second_cell.get("is_lab", False)):
        return False

    first_session_id = str(first_cell.get("session_id", "")).strip()
    second_session_id = str(second_cell.get("session_id", "")).strip()
    if not first_session_id or first_session_id != second_session_id:
        return False

    if int(first_cell.get("session_length", 0)) != 2 or int(second_cell.get("session_length", 0)) != 2:
        return False

    return {int(first_cell.get("session_part", 0)), int(second_cell.get("session_part", 0))} == {1, 2}


def _sanitize_sheet_name(name: str, index: int, prefix: str = "S") -> str:
    cleaned = name
    for illegal in ("[", "]", ":", "*", "?", "/", "\\"):
        cleaned = cleaned.replace(illegal, "-")

    sheet_prefix = f"{prefix}{index + 1}_"
    limit = 31 - len(sheet_prefix)
    return (sheet_prefix + cleaned[:limit]).strip() or f"{prefix}{index + 1}_Timetable"


def _normalize_lunch_period_number(period: int, periods_per_day: int) -> int:
    if period < 1 or period > periods_per_day:
        raise ValueError(
            f"Lunch break period {period} is out of range for {periods_per_day} periods/day."
        )
    return period


def _normalize_lunch_break_periods_by_day(
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None,
    working_days: int,
    periods_per_day: int,
) -> Dict[int, Set[int]]:
    normalized_by_day: Dict[int, Set[int]] = {
        day_index: set()
        for day_index in range(working_days)
    }

    if lunch_break_periods is None:
        return normalized_by_day

    if isinstance(lunch_break_periods, dict):
        for day_key, period_values in lunch_break_periods.items():
            day_index = int(day_key)
            if day_index < 0 or day_index >= working_days:
                raise ValueError(
                    f"Lunch break day index {day_index + 1} is out of range for {working_days} working days."
                )

            if isinstance(period_values, int):
                candidate_values = [period_values]
            else:
                candidate_values = list(period_values)

            day_set: Set[int] = set()
            for raw_period in candidate_values:
                day_set.add(_normalize_lunch_period_number(int(raw_period), periods_per_day))
            normalized_by_day[day_index] = day_set

        return normalized_by_day

    if (
        isinstance(lunch_break_periods, (list, tuple))
        and len(lunch_break_periods) == working_days
        and all(isinstance(item, int) for item in lunch_break_periods)
    ):
        for day_index, raw_period in enumerate(lunch_break_periods):
            normalized_by_day[day_index] = {
                _normalize_lunch_period_number(int(raw_period), periods_per_day)
            }
        return normalized_by_day

    shared_set: Set[int] = set()
    for raw_period in lunch_break_periods:
        shared_set.add(_normalize_lunch_period_number(int(raw_period), periods_per_day))

    for day_index in range(working_days):
        normalized_by_day[day_index] = set(shared_set)

    return normalized_by_day


def _normalize_lunch_break_periods(
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None,
    periods_per_day: int,
    working_days: int | None = None,
) -> Set[int]:
    if lunch_break_periods is None:
        return set()

    if isinstance(lunch_break_periods, dict):
        flattened: Set[int] = set()
        for period_values in lunch_break_periods.values():
            if isinstance(period_values, int):
                candidate_values = [period_values]
            else:
                candidate_values = list(period_values)

            for raw_period in candidate_values:
                flattened.add(_normalize_lunch_period_number(int(raw_period), periods_per_day))
        return flattened

    if working_days is not None and working_days > 0:
        by_day = _normalize_lunch_break_periods_by_day(
            lunch_break_periods,
            working_days,
            periods_per_day,
        )
        flattened: Set[int] = set()
        for period_set in by_day.values():
            flattened.update(period_set)
        return flattened

    normalized: Set[int] = set()
    for raw_period in lunch_break_periods:
        normalized.add(_normalize_lunch_period_number(int(raw_period), periods_per_day))
    return normalized


def _is_lunch_period(
    lunch_break_periods_by_day: Dict[int, Set[int]],
    day_index: int,
    period_number: int,
) -> bool:
    return period_number in lunch_break_periods_by_day.get(day_index, set())


def _build_multi_lecture_pool(assignments: List[FacultyAssignment]) -> List[Dict[str, object]]:
    lecture_pool: List[Dict[str, object]] = []
    session_id = 1

    for assignment in assignments:
        class_name = str(assignment["class_name"])
        standard = str(assignment["standard"])
        section = str(assignment["section"])
        _, group_number = _extract_section_group(section)
        section_key = build_section_key(class_name, standard, section)
        section_family_key = _build_section_family_key(class_name, standard, section)
        bundle_key = str(assignment.get("lab_bundle_key", "")).strip()
        lab_bundle_order_raw = assignment.get("lab_bundle_order")
        lab_subject_count_raw = assignment.get("lab_subject_count")
        lab_bundle_order = None
        if lab_bundle_order_raw is not None:
            try:
                lab_bundle_order = int(lab_bundle_order_raw)
            except (TypeError, ValueError):
                lab_bundle_order = None

        lab_subject_count = None
        if lab_subject_count_raw is not None:
            try:
                lab_subject_count = int(lab_subject_count_raw)
            except (TypeError, ValueError):
                lab_subject_count = None

        lecture_periods = int(assignment["lectures"])
        is_lab = bool(assignment.get("is_lab", False))
        duration = 2 if is_lab else 1

        if is_lab and lecture_periods % 2 != 0:
            raise ValueError(
                "Lab lectures must be in even counts because labs are scheduled in contiguous double periods."
            )

        session_count = lecture_periods // duration
        for _ in range(session_count):
            lecture_pool.append(
                {
                    "class_name": class_name,
                    "standard": standard,
                    "section": section,
                    "section_key": section_key,
                    "section_family_key": section_family_key,
                    "section_group_number": group_number,
                    "subject": str(assignment["subject"]),
                    "teacher": str(assignment["teacher"]),
                    "is_lab": is_lab,
                    "allowed_rooms": list(assignment.get("allowed_rooms", [])),
                    "subject_daily_max": assignment.get("subject_daily_max"),
                    "lab_bundle_key": bundle_key,
                    "lab_bundle_order": lab_bundle_order,
                    "lab_subject_count": lab_subject_count,
                    "duration": duration,
                    "session_id": session_id,
                }
            )
            session_id += 1

    return lecture_pool


def _expected_subject_counts(assignments: List[FacultyAssignment]) -> Dict[Tuple[str, str], int]:
    expected: Dict[Tuple[str, str], int] = {}
    for item in assignments:
        section_key = build_section_key(
            str(item["class_name"]),
            str(item["standard"]),
            str(item["section"]),
        )
        subject = str(item["subject"])
        key = (section_key, subject)
        expected[key] = expected.get(key, 0) + int(item["lectures"])
    return expected


def _resolve_candidate_rooms_for_lecture(
    lecture: Dict[str, object],
    rooms: List[str],
    section_home_rooms: Dict[str, str],
) -> List[str]:
    room_set = set(rooms)
    section_key = str(lecture["section_key"])
    is_lab = bool(lecture.get("is_lab", False))
    allowed_rooms = [str(room) for room in lecture.get("allowed_rooms", []) if str(room).strip()]

    if is_lab:
        if allowed_rooms:
            return [room for room in allowed_rooms if room in room_set]
        return list(rooms)

    home_room = section_home_rooms.get(section_key)
    if home_room:
        return [home_room] if home_room in room_set else []

    if allowed_rooms:
        constrained = [room for room in allowed_rooms if room in room_set]
        if constrained:
            return constrained

    return list(rooms)


def _build_section_tables_from_schedule(
    schedule_by_section: Dict[str, Dict[Tuple[int, int], TimetableCell | None]],
    section_metadata: Dict[str, Dict[str, str]],
    working_days: int,
    periods_per_day: int,
) -> List[SectionTable]:
    section_tables: List[SectionTable] = []

    for section_key in sorted(section_metadata.keys()):
        matrix: TimetableMatrix = []
        for day in range(working_days):
            row: List[TimetableCell | None] = []
            for period in range(periods_per_day):
                row.append(schedule_by_section[section_key][(day, period)])
            matrix.append(row)

        section_tables.append(
            {
                "section_key": section_key,
                "class_name": section_metadata[section_key]["class_name"],
                "standard": section_metadata[section_key]["standard"],
                "section": section_metadata[section_key]["section"],
                "timetable": matrix,
            }
        )

    return section_tables


def build_faculty_timetable_tables(
    days: List[str],
    periods: List[str],
    section_tables: List[SectionTable],
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> List[FacultyTable]:
    day_count = len(days)
    period_count = len(periods)
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        day_count,
        period_count,
    )

    faculty_matrix_map: Dict[str, TimetableMatrix] = {}

    for table in section_tables:
        matrix: TimetableMatrix = table["timetable"]
        for day_index in range(day_count):
            for period_index in range(period_count):
                if _is_lunch_period(lunch_by_day, day_index, period_index + 1):
                    continue

                section_cell = matrix[day_index][period_index]
                if section_cell is None:
                    continue

                teacher_name = str(section_cell.get("teacher", "")).strip()
                if not teacher_name:
                    continue

                if teacher_name not in faculty_matrix_map:
                    faculty_matrix_map[teacher_name] = [
                        [None for _ in range(period_count)]
                        for _ in range(day_count)
                    ]

                faculty_cell: TimetableCell = {
                    "subject": str(section_cell.get("subject", "")),
                    "class_name": str(table["class_name"]),
                    "standard": str(table["standard"]),
                    "section": str(table["section"]),
                    "section_key": str(table["section_key"]),
                    "room": str(section_cell.get("room", "")),
                    "is_lab": bool(section_cell.get("is_lab", False)),
                    "session_length": int(section_cell.get("session_length", 1)),
                    "session_part": int(section_cell.get("session_part", 1)),
                }

                existing_cell = faculty_matrix_map[teacher_name][day_index][period_index]
                if existing_cell is None:
                    faculty_matrix_map[teacher_name][day_index][period_index] = faculty_cell
                elif isinstance(existing_cell, dict) and "conflicts" in existing_cell:
                    existing_conflicts = list(existing_cell.get("conflicts", []))
                    existing_conflicts.append(faculty_cell)
                    faculty_matrix_map[teacher_name][day_index][period_index] = {
                        "conflicts": existing_conflicts
                    }
                else:
                    faculty_matrix_map[teacher_name][day_index][period_index] = {
                        "conflicts": [existing_cell, faculty_cell]
                    }

    return [
        {
            "teacher": teacher_name,
            "timetable": faculty_matrix_map[teacher_name],
        }
        for teacher_name in sorted(faculty_matrix_map.keys())
    ]


def build_class_timetable_tables(
    section_tables: List[SectionTable],
) -> List[ClassTable]:
    def merge_grouped_section_timetable(grouped_sections: List[Dict[str, object]]) -> TimetableMatrix:
        if not grouped_sections:
            return []

        first_matrix = grouped_sections[0]["timetable"]
        day_count = len(first_matrix)
        period_count = len(first_matrix[0]) if day_count > 0 else 0
        merged_matrix: TimetableMatrix = []

        for day_index in range(day_count):
            merged_row: List[TimetableCell | None] = []
            for period_index in range(period_count):
                slot_cells: List[Tuple[str, TimetableCell]] = []
                for section_entry in grouped_sections:
                    section_code = str(section_entry.get("section", ""))
                    section_matrix = section_entry["timetable"]
                    section_cell = section_matrix[day_index][period_index]
                    if section_cell is None:
                        continue
                    slot_cells.append((section_code, section_cell))

                if not slot_cells:
                    merged_row.append(None)
                    continue

                if len(slot_cells) == 1:
                    merged_row.append(slot_cells[0][1])
                    continue

                bundle_keys = {str(cell.get("lab_bundle_key", "")).strip() for _, cell in slot_cells}
                if (
                    len(slot_cells) == 3
                    and all(bool(cell.get("is_lab", False)) for _, cell in slot_cells)
                    and len(bundle_keys) == 1
                    and "" not in bundle_keys
                ):
                    sorted_groups = sorted(
                        slot_cells,
                        key=lambda item: (
                            _extract_section_group(item[0])[1] is None,
                            _extract_section_group(item[0])[1] or 0,
                            item[0],
                        ),
                    )

                    merged_cell = dict(sorted_groups[0][1])
                    subgroup_details: List[Dict[str, str]] = []
                    for section_code, subgroup_cell in sorted_groups:
                        _, group_number = _extract_section_group(section_code)
                        subgroup_details.append(
                            {
                                "group": f"G{group_number}" if group_number is not None else section_code,
                                "section": section_code,
                                "subject": str(subgroup_cell.get("subject", "")),
                                "teacher": str(subgroup_cell.get("teacher", "")),
                                "room": str(subgroup_cell.get("room", "")),
                            }
                        )

                    merged_cell["subject"] = "LAB Group Session"
                    merged_cell["teacher"] = "Multiple Faculty"
                    merged_cell["room"] = "Multiple Labs"
                    merged_cell["lab_group_details"] = subgroup_details
                    merged_cell["group_count"] = len(subgroup_details)
                    merged_row.append(merged_cell)
                    continue

                # Fallback to the base section cell when mixed entries appear in the same slot.
                fallback_cell = sorted(
                    slot_cells,
                    key=lambda item: (
                        _extract_section_group(item[0])[1] is not None,
                        _extract_section_group(item[0])[1] or 0,
                        item[0],
                    ),
                )[0][1]
                merged_row.append(fallback_cell)

            merged_matrix.append(merged_row)

        return merged_matrix

    grouped: Dict[str, Dict[str, object]] = {}

    for table in section_tables:
        class_name = str(table["class_name"])
        standard = str(table["standard"])
        class_key = build_class_key(class_name, standard)
        base_section, group_number = _extract_section_group(str(table["section"]))

        if class_key not in grouped:
            grouped[class_key] = {
                "class_key": class_key,
                "class_name": class_name,
                "standard": standard,
                "sections_by_base": {},
            }

        sections_by_base = grouped[class_key]["sections_by_base"]
        if base_section not in sections_by_base:
            sections_by_base[base_section] = []

        sections_by_base[base_section].append(
            {
                "section": str(table["section"]),
                "section_key": str(table["section_key"]),
                "group_number": group_number,
                "timetable": table["timetable"],
            }
        )

    class_tables: List[ClassTable] = []
    for class_key in sorted(grouped.keys()):
        class_entry = grouped[class_key]
        sections_by_base = class_entry.get("sections_by_base", {})
        merged_sections: List[Dict[str, object]] = []

        for base_section in sorted(sections_by_base.keys()):
            grouped_sections = sorted(
                sections_by_base[base_section],
                key=lambda item: (
                    item.get("group_number") is not None,
                    int(item["group_number"]) if item.get("group_number") is not None else 0,
                    str(item.get("section", "")),
                ),
            )
            anchor_entry = grouped_sections[0]
            merged_sections.append(
                {
                    "section": base_section,
                    "section_key": str(anchor_entry.get("section_key", "")),
                    "timetable": merge_grouped_section_timetable(grouped_sections),
                }
            )

        class_tables.append(
            {
                "class_key": class_key,
                "class_name": str(class_entry["class_name"]),
                "standard": str(class_entry["standard"]),
                "sections": merged_sections,
            }
        )

    return class_tables


def build_room_timetable_tables(
    days: List[str],
    periods: List[str],
    section_tables: List[SectionTable],
    room_names: List[str] | None = None,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> List[RoomTable]:
    day_count = len(days)
    period_count = len(periods)
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        day_count,
        period_count,
    )

    room_matrix_map: Dict[str, TimetableMatrix] = {}
    ordered_room_names = sorted({str(room) for room in (room_names or []) if str(room).strip()})

    for room_name in ordered_room_names:
        room_matrix_map[room_name] = [
            [None for _ in range(period_count)]
            for _ in range(day_count)
        ]

    for table in section_tables:
        matrix: TimetableMatrix = table["timetable"]
        for day_index in range(day_count):
            for period_index in range(period_count):
                if _is_lunch_period(lunch_by_day, day_index, period_index + 1):
                    continue

                section_cell = matrix[day_index][period_index]
                if section_cell is None:
                    continue

                room_name = str(section_cell.get("room", "")).strip()
                if not room_name:
                    continue

                if room_name not in room_matrix_map:
                    room_matrix_map[room_name] = [
                        [None for _ in range(period_count)]
                        for _ in range(day_count)
                    ]

                room_cell: TimetableCell = {
                    "subject": str(section_cell.get("subject", "")),
                    "teacher": str(section_cell.get("teacher", "")),
                    "class_name": str(table["class_name"]),
                    "standard": str(table["standard"]),
                    "section": str(table["section"]),
                    "section_key": str(table["section_key"]),
                    "is_lab": bool(section_cell.get("is_lab", False)),
                    "session_length": int(section_cell.get("session_length", 1)),
                    "session_part": int(section_cell.get("session_part", 1)),
                }

                existing_cell = room_matrix_map[room_name][day_index][period_index]
                if existing_cell is None:
                    room_matrix_map[room_name][day_index][period_index] = room_cell
                elif isinstance(existing_cell, dict) and "conflicts" in existing_cell:
                    existing_conflicts = list(existing_cell.get("conflicts", []))
                    existing_conflicts.append(room_cell)
                    room_matrix_map[room_name][day_index][period_index] = {
                        "conflicts": existing_conflicts
                    }
                else:
                    room_matrix_map[room_name][day_index][period_index] = {
                        "conflicts": [existing_cell, room_cell]
                    }

    return [
        {
            "room": room_name,
            "timetable": room_matrix_map[room_name],
        }
        for room_name in sorted(room_matrix_map.keys())
    ]


def _calculate_soft_score(
    section_tables: List[SectionTable],
    working_days: int,
    periods_per_day: int,
    lunch_break_periods_by_day: Dict[int, Set[int]],
) -> Tuple[float, Dict[str, float]]:
    teaching_periods_by_day: Dict[int, List[int]] = {
        day_index: [
            period
            for period in range(periods_per_day)
            if not _is_lunch_period(lunch_break_periods_by_day, day_index, period + 1)
        ]
        for day_index in range(working_days)
    }

    if not any(teaching_periods_by_day.values()):
        return 0.0, {
            "gap_penalty": 0.0,
            "pre_lunch_free_penalty": 0.0,
            "subject_repeat_penalty": 0.0,
            "teacher_balance_penalty": 0.0,
            "section_balance_penalty": 0.0,
            "lab_edge_penalty": 0.0,
            "total_penalty": 0.0,
        }

    gap_penalty = 0.0
    pre_lunch_free_penalty = 0.0
    subject_repeat_penalty = 0.0
    teacher_balance_penalty = 0.0
    section_balance_penalty = 0.0
    lab_edge_penalty = 0.0

    teacher_day_load: Dict[Tuple[str, int], int] = {}

    for table in section_tables:
        matrix: TimetableMatrix = table["timetable"]
        section_daily_loads: List[int] = []

        for day in range(working_days):
            row = matrix[day]
            teaching_periods = teaching_periods_by_day.get(day, [])
            rank_by_period = {period: rank for rank, period in enumerate(teaching_periods)}
            occupied = [period for period in teaching_periods if row[period] is not None]
            section_daily_loads.append(len(occupied))

            day_lunch_periods = sorted(lunch_break_periods_by_day.get(day, set()))
            if day_lunch_periods:
                first_lunch_period_index = day_lunch_periods[0] - 1
                pre_lunch_periods = [period for period in teaching_periods if period < first_lunch_period_index]
                if pre_lunch_periods:
                    free_before_lunch = sum(1 for period in pre_lunch_periods if row[period] is None)
                    pre_lunch_free_penalty += free_before_lunch * 0.05

            if occupied:
                first_period = min(occupied)
                last_period = max(occupied)
                gap_count = sum(
                    1
                    for period in teaching_periods
                    if first_period < period < last_period and row[period] is None
                )
                gap_penalty += gap_count * 2.0

            subject_counter: Counter[str] = Counter()
            for period in occupied:
                cell = row[period]
                if cell is None:
                    continue

                subject_counter[str(cell["subject"])] += 1
                teacher_name = str(cell["teacher"])
                teacher_day_key = (teacher_name, day)
                teacher_day_load[teacher_day_key] = teacher_day_load.get(teacher_day_key, 0) + 1

                if bool(cell.get("is_lab", False)) and int(cell.get("session_part", 1)) == 1:
                    period_rank = rank_by_period[period]
                    if period_rank == 0 or period_rank >= max(0, len(teaching_periods) - 2):
                        lab_edge_penalty += 0.6

            repeated_subjects = sum(max(0, count - 1) for count in subject_counter.values())
            subject_repeat_penalty += repeated_subjects * 1.2

        section_mean = sum(section_daily_loads) / len(section_daily_loads) if section_daily_loads else 0.0
        section_variance = (
            sum((load - section_mean) ** 2 for load in section_daily_loads) / len(section_daily_loads)
            if section_daily_loads
            else 0.0
        )
        section_balance_penalty += section_variance * 0.8

    teacher_names = sorted({teacher for teacher, _ in teacher_day_load})
    for teacher_name in teacher_names:
        loads = [teacher_day_load.get((teacher_name, day), 0) for day in range(working_days)]
        teacher_mean = sum(loads) / len(loads) if loads else 0.0
        teacher_variance = (
            sum((load - teacher_mean) ** 2 for load in loads) / len(loads)
            if loads
            else 0.0
        )
        teacher_balance_penalty += teacher_variance * 1.1

    total_penalty = (
        gap_penalty
        + pre_lunch_free_penalty
        + subject_repeat_penalty
        + teacher_balance_penalty
        + section_balance_penalty
        + lab_edge_penalty
    )
    score = max(0.0, round(100.0 - total_penalty, 2))

    return score, {
        "gap_penalty": round(gap_penalty, 2),
        "pre_lunch_free_penalty": round(pre_lunch_free_penalty, 2),
        "subject_repeat_penalty": round(subject_repeat_penalty, 2),
        "teacher_balance_penalty": round(teacher_balance_penalty, 2),
        "section_balance_penalty": round(section_balance_penalty, 2),
        "lab_edge_penalty": round(lab_edge_penalty, 2),
        "total_penalty": round(total_penalty, 2),
    }


def generate_multi_section_timetable(
    assignments: List[FacultyAssignment],
    rooms: List[str],
    working_days: int,
    periods_per_day: int = 6,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
    section_home_rooms: Dict[str, str] | None = None,
    faculty_daily_default_max: int | None = None,
    faculty_daily_limits: Dict[str, int] | None = None,
    faculty_unavailability: Dict[str, Set[Tuple[int, int]]] | None = None,
    faculty_preferences: Dict[str, Dict[str, object]] | None = None,
    fixed_slot_constraints: List[Dict[str, object]] | None = None,
    locked_section_cells: List[Dict[str, object]] | None = None,
    require_daily_free_period: bool = False,
    lab_same_day_subject_threshold: int = 3,
    max_retries: int = 3000,
    random_seed: int | None = None,
) -> Tuple[List[str], List[str], List[SectionTable], Dict[str, object]]:
    """Generate conflict-free timetables and optimize candidate quality using soft constraints."""
    if working_days < 1 or working_days > len(DAY_NAMES):
        raise ValueError("Working days must be between 1 and 7.")

    if not assignments:
        raise ValueError("At least one faculty assignment is required.")

    if not rooms:
        raise ValueError("At least one room is required.")

    if int(lab_same_day_subject_threshold) < 1:
        raise ValueError("LAB same-day threshold must be at least 1.")

    assignments = _expand_assignments_for_lab_groups(assignments)

    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        working_days,
        periods_per_day,
    )
    section_home_rooms = section_home_rooms or {}
    faculty_daily_limits = faculty_daily_limits or {}
    faculty_unavailability = faculty_unavailability or {}
    faculty_preferences = faculty_preferences or {}
    fixed_slot_constraints = fixed_slot_constraints or []
    locked_section_cells = locked_section_cells or []

    slots = create_time_slots(working_days, periods_per_day)
    assignable_slots = [
        slot
        for slot in slots
        if not _is_lunch_period(lunch_by_day, slot[0], slot[1] + 1)
    ]
    assignable_slot_set = set(assignable_slots)

    if not assignable_slots:
        raise ValueError("No teaching slots available after applying lunch break periods.")

    normalized_unavailability: Dict[str, Set[Tuple[int, int]]] = {}
    for teacher_name, slot_items in faculty_unavailability.items():
        teacher_key = str(teacher_name).strip()
        if not teacher_key:
            continue

        teacher_slots: Set[Tuple[int, int]] = set()
        for raw_slot in slot_items:
            if not isinstance(raw_slot, tuple) or len(raw_slot) != 2:
                raise ValueError(
                    f"Invalid unavailable slot configured for faculty '{teacher_key}'."
                )

            day_index = int(raw_slot[0])
            period_index = int(raw_slot[1])
            slot = (day_index, period_index)
            if slot not in assignable_slot_set:
                raise ValueError(
                    f"Unavailable slot for faculty '{teacher_key}' is outside teaching slots: "
                    f"day {day_index + 1}, period {period_index + 1}."
                )
            teacher_slots.add(slot)

        if teacher_slots:
            normalized_unavailability[teacher_key] = teacher_slots

    normalized_preferences: Dict[str, Dict[str, object]] = {}
    for teacher_name, raw_config in faculty_preferences.items():
        teacher_key = str(teacher_name).strip()
        if not teacher_key:
            continue

        config = raw_config if isinstance(raw_config, dict) else {}
        preferred_slots_raw = config.get("preferred_slots", set())
        avoid_slots_raw = config.get("avoid_slots", set())
        max_consecutive_raw = config.get("max_consecutive")

        preferred_slots: Set[Tuple[int, int]] = set()
        avoid_slots: Set[Tuple[int, int]] = set()

        for raw_slot in preferred_slots_raw:
            if not isinstance(raw_slot, tuple) or len(raw_slot) != 2:
                raise ValueError(
                    f"Preferred slot configuration for faculty '{teacher_key}' is invalid."
                )
            slot = (int(raw_slot[0]), int(raw_slot[1]))
            if slot not in assignable_slot_set:
                raise ValueError(
                    f"Preferred slot for faculty '{teacher_key}' is outside teaching slots: "
                    f"day {slot[0] + 1}, period {slot[1] + 1}."
                )
            preferred_slots.add(slot)

        for raw_slot in avoid_slots_raw:
            if not isinstance(raw_slot, tuple) or len(raw_slot) != 2:
                raise ValueError(
                    f"Avoid slot configuration for faculty '{teacher_key}' is invalid."
                )
            slot = (int(raw_slot[0]), int(raw_slot[1]))
            if slot not in assignable_slot_set:
                raise ValueError(
                    f"Avoid slot for faculty '{teacher_key}' is outside teaching slots: "
                    f"day {slot[0] + 1}, period {slot[1] + 1}."
                )
            avoid_slots.add(slot)

        overlap_slots = preferred_slots & avoid_slots
        if overlap_slots:
            raise ValueError(
                f"Faculty '{teacher_key}' has overlapping preferred and avoid slots."
            )

        max_consecutive = None
        if max_consecutive_raw is not None:
            max_consecutive = int(max_consecutive_raw)
            if max_consecutive < 1:
                raise ValueError(
                    f"Max consecutive classes for faculty '{teacher_key}' must be at least 1."
                )

        normalized_preferences[teacher_key] = {
            "preferred_slots": preferred_slots,
            "avoid_slots": avoid_slots,
            "max_consecutive": max_consecutive,
        }

    normalized_fixed_slots: List[Dict[str, object]] = []
    for index, raw_constraint in enumerate(fixed_slot_constraints, start=1):
        try:
            section_key = str(raw_constraint["section_key"])
            subject_name = str(raw_constraint["subject"])
            teacher_name = str(raw_constraint["teacher"])
            day_index = int(raw_constraint["day_index"])
            start_period = int(raw_constraint["start_period"])
        except (KeyError, TypeError, ValueError) as error:
            raise ValueError(
                f"Fixed-slot constraint #{index} is missing required fields."
            ) from error

        if day_index < 0 or day_index >= working_days:
            raise ValueError(
                f"Fixed-slot constraint #{index} has invalid day index {day_index + 1}."
            )

        if start_period < 0 or start_period >= periods_per_day:
            raise ValueError(
                f"Fixed-slot constraint #{index} has invalid start period {start_period + 1}."
            )

        duration = int(raw_constraint.get("duration", 1) or 1)
        if duration < 1:
            raise ValueError(
                f"Fixed-slot constraint #{index} has invalid duration {duration}."
            )

        if start_period + duration > periods_per_day:
            raise ValueError(
                f"Fixed-slot constraint #{index} does not fit in day timetable window."
            )

        block_slots = [(day_index, start_period + offset) for offset in range(duration)]
        if any(slot not in assignable_slot_set for slot in block_slots):
            raise ValueError(
                f"Fixed-slot constraint #{index} intersects lunch/non-teaching periods."
            )

        room_name = str(raw_constraint.get("room", "")).strip()
        if room_name and room_name not in rooms:
            raise ValueError(
                f"Fixed-slot constraint #{index} references unknown room '{room_name}'."
            )

        normalized_fixed_slots.append(
            {
                "section_key": section_key,
                "subject": subject_name,
                "teacher": teacher_name,
                "day_index": day_index,
                "start_period": start_period,
                "duration": duration,
                "room": room_name,
            }
        )

    normalized_fixed_slots.sort(
        key=lambda item: (
            int(item["day_index"]),
            int(item["start_period"]),
            str(item["section_key"]),
            str(item["subject"]),
            str(item["teacher"]),
        )
    )

    section_fixed_slot_usage: Set[Tuple[str, int, int]] = set()
    teacher_fixed_slot_usage: Set[Tuple[str, int, int]] = set()
    room_fixed_slot_usage: Set[Tuple[str, int, int]] = set()

    for index, fixed_slot in enumerate(normalized_fixed_slots, start=1):
        section_key = str(fixed_slot["section_key"])
        teacher_name = str(fixed_slot["teacher"])
        room_name = str(fixed_slot.get("room", ""))
        day_index = int(fixed_slot["day_index"])
        start_period = int(fixed_slot["start_period"])
        duration = int(fixed_slot["duration"])
        block_slots = [(day_index, start_period + offset) for offset in range(duration)]

        preference_config = normalized_preferences.get(teacher_name, {})
        avoid_slots = set(preference_config.get("avoid_slots", set()))
        if any(slot in avoid_slots for slot in block_slots):
            raise ValueError(
                f"Fixed-slot constraint #{index} places faculty '{teacher_name}' in an avoided slot."
            )

        unavailable_slots = normalized_unavailability.get(teacher_name, set())
        if any(slot in unavailable_slots for slot in block_slots):
            raise ValueError(
                f"Fixed-slot constraint #{index} assigns faculty '{teacher_name}' "
                "during an unavailable slot."
            )

        for day_slot in block_slots:
            section_slot_key = (section_key, day_slot[0], day_slot[1])
            if section_slot_key in section_fixed_slot_usage:
                raise ValueError(
                    f"Fixed-slot conflict detected for section {section_key} on day "
                    f"{day_slot[0] + 1}, period {day_slot[1] + 1}."
                )
            section_fixed_slot_usage.add(section_slot_key)

            teacher_slot_key = (teacher_name, day_slot[0], day_slot[1])
            if teacher_slot_key in teacher_fixed_slot_usage:
                raise ValueError(
                    f"Fixed-slot conflict detected for faculty '{teacher_name}' on day "
                    f"{day_slot[0] + 1}, period {day_slot[1] + 1}."
                )
            teacher_fixed_slot_usage.add(teacher_slot_key)

            if room_name:
                room_slot_key = (room_name, day_slot[0], day_slot[1])
                if room_slot_key in room_fixed_slot_usage:
                    raise ValueError(
                        f"Fixed-slot conflict detected for room '{room_name}' on day "
                        f"{day_slot[0] + 1}, period {day_slot[1] + 1}."
                    )
                room_fixed_slot_usage.add(room_slot_key)

    for mapped_room in section_home_rooms.values():
        if mapped_room not in rooms:
            raise ValueError(
                f"Configured section classroom '{mapped_room}' is not present in room list."
            )

    lecture_pool = _build_multi_lecture_pool(assignments)

    lecture_pool_count: Dict[Tuple[str, str, str, int], int] = {}
    for lecture in lecture_pool:
        key = (
            str(lecture["section_key"]),
            str(lecture["subject"]),
            str(lecture["teacher"]),
            int(lecture["duration"]),
        )
        lecture_pool_count[key] = lecture_pool_count.get(key, 0) + 1

    fixed_slot_count: Dict[Tuple[str, str, str, int], int] = {}
    for index, fixed_slot in enumerate(normalized_fixed_slots, start=1):
        key = (
            str(fixed_slot["section_key"]),
            str(fixed_slot["subject"]),
            str(fixed_slot["teacher"]),
            int(fixed_slot["duration"]),
        )
        fixed_slot_count[key] = fixed_slot_count.get(key, 0) + 1

    section_metadata: Dict[str, Dict[str, str]] = {}
    section_total_lectures: Dict[str, int] = {}
    section_subject_totals: Dict[Tuple[str, str], int] = {}
    section_subject_daily_overrides: Dict[Tuple[str, str], int] = {}
    section_subject_is_lab: Dict[Tuple[str, str], bool] = {}
    teacher_names: Set[str] = set()
    teacher_has_theory_load: Dict[str, bool] = {}

    for item in assignments:
        class_name = str(item["class_name"])
        standard = str(item["standard"])
        section = str(item["section"])
        section_key = build_section_key(class_name, standard, section)
        teacher_name = str(item["teacher"])
        subject_name = str(item["subject"])

        section_metadata[section_key] = {
            "class_name": class_name,
            "standard": standard,
            "section": section,
        }

        section_total_lectures[section_key] = (
            section_total_lectures.get(section_key, 0) + int(item["lectures"])
        )

        subject_key = (section_key, subject_name)
        section_subject_totals[subject_key] = (
            section_subject_totals.get(subject_key, 0) + int(item["lectures"])
        )
        section_subject_is_lab[subject_key] = section_subject_is_lab.get(subject_key, False) or bool(
            item.get("is_lab", False)
        )

        subject_daily_max = item.get("subject_daily_max")
        if subject_daily_max is not None:
            if subject_key in section_subject_daily_overrides and section_subject_daily_overrides[subject_key] != int(subject_daily_max):
                raise ValueError(
                    f"Conflicting subject daily max values provided for {section_key} -> {subject_name}."
                )
            section_subject_daily_overrides[subject_key] = int(subject_daily_max)

        teacher_names.add(teacher_name)
        if not bool(item.get("is_lab", False)):
            teacher_has_theory_load[teacher_name] = True
        else:
            teacher_has_theory_load.setdefault(teacher_name, False)

    normalized_locked_slots: List[Dict[str, object]] = []
    locked_slot_count: Dict[Tuple[str, str, str, int], int] = {}
    locked_section_slot_usage: Set[Tuple[str, int, int]] = set()
    locked_lab_markers: Set[Tuple[str, str]] = set()

    for index, raw_locked_slot in enumerate(locked_section_cells, start=1):
        if not isinstance(raw_locked_slot, dict):
            raise ValueError(f"Locked slot #{index} must be an object.")

        try:
            section_key = str(raw_locked_slot["section_key"])
            day_index = int(raw_locked_slot["day_index"])
            period_index = int(raw_locked_slot["period_index"])
            subject_name = str(raw_locked_slot["subject"])
            teacher_name = str(raw_locked_slot["teacher"])
            room_name = str(raw_locked_slot["room"])
        except (KeyError, TypeError, ValueError) as error:
            raise ValueError(
                f"Locked slot #{index} is missing required values."
            ) from error

        if section_key not in section_metadata:
            raise ValueError(
                f"Locked slot #{index} references unknown section {section_key}."
            )

        if day_index < 0 or day_index >= working_days:
            raise ValueError(
                f"Locked slot #{index} has invalid day index {day_index + 1}."
            )

        if period_index < 0 or period_index >= periods_per_day:
            raise ValueError(
                f"Locked slot #{index} has invalid period index {period_index + 1}."
            )

        if (day_index, period_index) not in assignable_slot_set:
            raise ValueError(
                f"Locked slot #{index} intersects lunch/non-teaching periods."
            )

        is_lab = bool(raw_locked_slot.get("is_lab", False))
        session_length = int(raw_locked_slot.get("session_length", 2 if is_lab else 1) or 1)
        session_part = int(raw_locked_slot.get("session_part", 1) or 1)
        if is_lab:
            session_length = max(2, session_length)
            if session_part not in {1, 2}:
                session_part = 1
        else:
            session_length = 1
            session_part = 1

        start_period = period_index if not (is_lab and session_part == 2) else period_index - 1
        if start_period < 0 or start_period + session_length > periods_per_day:
            raise ValueError(
                f"Locked slot #{index} does not fit within timetable day bounds."
            )

        block_slots = [(day_index, start_period + offset) for offset in range(session_length)]
        if any(slot not in assignable_slot_set for slot in block_slots):
            raise ValueError(
                f"Locked slot #{index} intersects lunch/non-teaching periods."
            )

        session_id = str(raw_locked_slot.get("session_id", "")).strip()
        if is_lab:
            marker = session_id or f"{section_key}:{subject_name}:{teacher_name}:{day_index}:{start_period}"
            marker_key = (section_key, marker)
            if marker_key in locked_lab_markers:
                continue
            locked_lab_markers.add(marker_key)

        for slot in block_slots:
            usage_key = (section_key, slot[0], slot[1])
            if usage_key in locked_section_slot_usage:
                raise ValueError(
                    f"Locked slot #{index} overlaps another locked slot in {section_key} on day {slot[0] + 1}, period {slot[1] + 1}."
                )
            locked_section_slot_usage.add(usage_key)

        section_info = section_metadata[section_key]
        normalized_locked_slots.append(
            {
                "section_key": section_key,
                "subject": subject_name,
                "teacher": teacher_name,
                "day_index": day_index,
                "start_period": start_period,
                "duration": session_length,
                "room": room_name,
                "is_locked": True,
                "class_name": section_info["class_name"],
                "standard": section_info["standard"],
                "section": section_info["section"],
            }
        )

        count_key = (section_key, subject_name, teacher_name, session_length)
        locked_slot_count[count_key] = locked_slot_count.get(count_key, 0) + 1

    for key, fixed_count in fixed_slot_count.items():
        available_count = lecture_pool_count.get(key, 0)
        locked_count = locked_slot_count.get(key, 0)
        if fixed_count + locked_count > available_count:
            raise ValueError(
                "Fixed or locked slots exceed available lecture sessions for "
                f"{key[0]} -> {key[1]} ({key[2]})."
            )

    for key, locked_count in locked_slot_count.items():
        if locked_count > lecture_pool_count.get(key, 0):
            raise ValueError(
                "Locked slots exceed available lecture sessions for "
                f"{key[0]} -> {key[1]} ({key[2]})."
            )

    preassigned_section_usage: Set[Tuple[str, int, int]] = set()
    preassigned_teacher_usage: Set[Tuple[str, int, int]] = set()
    preassigned_room_usage: Set[Tuple[str, int, int]] = set()

    for block in list(normalized_locked_slots) + list(normalized_fixed_slots):
        section_key = str(block["section_key"])
        teacher_name = str(block["teacher"])
        room_name = str(block.get("room", "")).strip()
        day_index = int(block["day_index"])
        start_period = int(block["start_period"])
        duration = int(block["duration"])

        for offset in range(duration):
            period_index = start_period + offset

            section_key_slot = (section_key, day_index, period_index)
            if section_key_slot in preassigned_section_usage:
                raise ValueError(
                    f"Locked/fixed preassignment overlap in {section_key} on day {day_index + 1}, period {period_index + 1}."
                )
            preassigned_section_usage.add(section_key_slot)

            teacher_slot = (teacher_name, day_index, period_index)
            if teacher_slot in preassigned_teacher_usage:
                raise ValueError(
                    f"Locked/fixed preassignment creates faculty clash for {teacher_name} on day {day_index + 1}, period {period_index + 1}."
                )
            preassigned_teacher_usage.add(teacher_slot)

            if room_name:
                room_slot = (room_name, day_index, period_index)
                if room_slot in preassigned_room_usage:
                    raise ValueError(
                        f"Locked/fixed preassignment creates room clash for {room_name} on day {day_index + 1}, period {period_index + 1}."
                    )
                preassigned_room_usage.add(room_slot)

    slot_count_per_section = len(assignable_slots)
    for section_key, lecture_count in section_total_lectures.items():
        if lecture_count > slot_count_per_section:
            raise ValueError(
                f"{section_key} has {lecture_count} lectures but only {slot_count_per_section} available teaching slots."
            )

    total_lecture_periods = sum(int(item["lectures"]) for item in assignments)
    max_parallel_per_slot = min(len(section_metadata), len(rooms))
    max_global_capacity = slot_count_per_section * max_parallel_per_slot
    if total_lecture_periods > max_global_capacity:
        raise ValueError(
            "Total lectures exceed global capacity across sections. "
            "Increase rooms/periods/days or reduce lecture load."
        )

    max_teaching_periods_per_day = min(
        periods_per_day - len(lunch_by_day.get(day_index, set()))
        for day_index in range(working_days)
    )
    if max_teaching_periods_per_day <= 0:
        raise ValueError("Faculty max periods cannot be applied because there are no teaching periods per day.")

    for config in normalized_preferences.values():
        max_consecutive = config.get("max_consecutive")
        if max_consecutive is None:
            continue
        config["max_consecutive"] = min(int(max_consecutive), max_teaching_periods_per_day)

    daily_subject_limits: Dict[Tuple[str, str], int] = {}
    for subject_key, lecture_count in section_subject_totals.items():
        default_limit = max(1, math.ceil(lecture_count / working_days))
        if section_subject_is_lab.get(subject_key, False):
            default_limit = max(2, default_limit)

        chosen_limit = section_subject_daily_overrides.get(subject_key, default_limit)
        if section_subject_is_lab.get(subject_key, False):
            chosen_limit = max(2, min(chosen_limit, max_teaching_periods_per_day))
        else:
            chosen_limit = max(1, min(chosen_limit, max_teaching_periods_per_day))
        daily_subject_limits[subject_key] = chosen_limit

    for item in assignments:
        if bool(item.get("is_lab", False)):
            subject_key = (
                build_section_key(
                    str(item["class_name"]),
                    str(item["standard"]),
                    str(item["section"]),
                ),
                str(item["subject"]),
            )
            if daily_subject_limits[subject_key] < 2:
                raise ValueError(
                    f"Subject daily max for {subject_key[0]} -> {subject_key[1]} must be at least 2 for double-period labs."
                )

    default_limit = (
        max_teaching_periods_per_day
        if faculty_daily_default_max is None
        else max(1, min(faculty_daily_default_max, max_teaching_periods_per_day))
    )
    teacher_daily_limit: Dict[str, int] = {}
    for teacher_name in teacher_names:
        custom_limit = faculty_daily_limits.get(teacher_name, default_limit)
        if custom_limit < 1:
            raise ValueError(f"Daily max periods for faculty '{teacher_name}' must be at least 1.")

        effective_limit = min(custom_limit, max_teaching_periods_per_day)
        if (
            require_daily_free_period
            and max_teaching_periods_per_day > 1
            and teacher_has_theory_load.get(teacher_name, False)
        ):
            effective_limit = min(effective_limit, max_teaching_periods_per_day - 1)

        teacher_daily_limit[teacher_name] = max(1, effective_limit)

    fixed_teacher_day_load: Dict[Tuple[str, int], int] = {}
    for fixed_slot in normalized_fixed_slots:
        teacher_name = str(fixed_slot["teacher"])
        day_index = int(fixed_slot["day_index"])
        duration = int(fixed_slot["duration"])
        teacher_day_key = (teacher_name, day_index)
        fixed_teacher_day_load[teacher_day_key] = fixed_teacher_day_load.get(teacher_day_key, 0) + duration

    for (teacher_name, day_index), fixed_load in fixed_teacher_day_load.items():
        max_allowed = teacher_daily_limit.get(teacher_name, default_limit)
        if fixed_load > max_allowed:
            raise ValueError(
                f"Fixed-slot assignments exceed daily max for faculty '{teacher_name}' on day "
                f"{day_index + 1}: {fixed_load}>{max_allowed}."
            )

    durations = sorted({int(lecture["duration"]) for lecture in lecture_pool})
    start_slots_by_duration: Dict[int, List[Tuple[int, int]]] = {}
    for duration in durations:
        candidate_starts: List[Tuple[int, int]] = []
        for day in range(working_days):
            for start_period in range(0, periods_per_day - duration + 1):
                block_slots = [(day, start_period + offset) for offset in range(duration)]
                if all(slot in assignable_slot_set for slot in block_slots):
                    candidate_starts.append((day, start_period))

        if not candidate_starts:
            raise ValueError(
                f"No valid contiguous slot window available for duration {duration}."
            )
        start_slots_by_duration[duration] = candidate_starts

    first_lunch_period_index_by_day: Dict[int, int | None] = {
        day_index: (
            min(lunch_by_day.get(day_index, set())) - 1
            if lunch_by_day.get(day_index, set())
            else None
        )
        for day_index in range(working_days)
    }

    def preference_hit_counts(teacher_name: str, block_slots: List[Tuple[int, int]]) -> Tuple[int, int]:
        config = normalized_preferences.get(teacher_name, {})
        preferred_slots = set(config.get("preferred_slots", set()))
        avoid_slots = set(config.get("avoid_slots", set()))
        preferred_hits = sum(1 for slot in block_slots if slot in preferred_slots)
        avoid_hits = sum(1 for slot in block_slots if slot in avoid_slots)
        return preferred_hits, avoid_hits

    def would_exceed_max_consecutive(
        teacher_name: str,
        day_index: int,
        block_slots: List[Tuple[int, int]],
        teacher_usage_by_slot: Dict[Tuple[int, int], Set[str]],
    ) -> bool:
        config = normalized_preferences.get(teacher_name, {})
        max_consecutive = config.get("max_consecutive")
        if max_consecutive is None:
            return False

        occupied_periods: Set[int] = {
            slot[1]
            for slot in assignable_slots
            if slot[0] == day_index and teacher_name in teacher_usage_by_slot[slot]
        }
        occupied_periods.update(slot[1] for slot in block_slots)

        run_length = 0
        previous_period = None
        for period_index in sorted(occupied_periods):
            if previous_period is None or period_index == previous_period + 1:
                run_length += 1
            else:
                run_length = 1

            if run_length > int(max_consecutive):
                return True
            previous_period = period_index

        return False

    def would_violate_faculty_rest_gap(
        teacher_name: str,
        block_slots: List[Tuple[int, int]],
        teacher_usage_by_slot: Dict[Tuple[int, int], Set[str]],
    ) -> bool:
        if not block_slots:
            return False

        day_index = block_slots[0][0]
        first_period = min(slot[1] for slot in block_slots)
        last_period = max(slot[1] for slot in block_slots)

        previous_slot = (day_index, first_period - 1)
        if first_period - 1 >= 0 and teacher_name in teacher_usage_by_slot.get(previous_slot, set()):
            return True

        next_slot = (day_index, last_period + 1)
        if last_period + 1 < periods_per_day and teacher_name in teacher_usage_by_slot.get(next_slot, set()):
            return True

        return False

    def choose_distinct_rooms_for_bundle(room_options: List[List[str]]) -> List[str] | None:
        if not room_options:
            return []

        member_order = sorted(range(len(room_options)), key=lambda member_index: len(room_options[member_index]))
        assigned_rooms: Dict[int, str] = {}
        used_rooms: Set[str] = set()

        def backtrack(order_index: int) -> bool:
            if order_index >= len(member_order):
                return True

            member_index = member_order[order_index]
            candidate_rooms = room_options[member_index]
            for room_name in rng.sample(candidate_rooms, len(candidate_rooms)):
                if room_name in used_rooms:
                    continue

                used_rooms.add(room_name)
                assigned_rooms[member_index] = room_name
                if backtrack(order_index + 1):
                    return True

                used_rooms.remove(room_name)
                assigned_rooms.pop(member_index, None)

            return False

        if not backtrack(0):
            return None

        return [assigned_rooms[member_index] for member_index in range(len(room_options))]

    rng = random.Random(random_seed)
    day_labels = DAY_NAMES[:working_days]
    period_labels = [f"Period {index}" for index in range(1, periods_per_day + 1)]

    best_section_tables: List[SectionTable] | None = None
    best_score = -1.0
    best_penalties: Dict[str, float] = {}
    successful_candidates = 0
    stagnant_successes = 0

    lecture_count = len(lecture_pool)
    if lecture_count >= 500:
        effective_retry_budget = min(max_retries, 350)
        target_successful_candidates = 15
        stagnation_limit = 6
    elif lecture_count >= 250:
        effective_retry_budget = min(max_retries, 700)
        target_successful_candidates = 30
        stagnation_limit = 10
    else:
        effective_retry_budget = max_retries
        target_successful_candidates = 120
        stagnation_limit = 40

    attempts_used = 0

    for attempt in range(effective_retry_budget):
        attempts_used = attempt + 1
        schedule_by_section: Dict[str, Dict[Tuple[int, int], TimetableCell | None]] = {
            section_key: {slot: None for slot in slots}
            for section_key in section_metadata
        }
        teacher_usage_by_slot = {slot: set() for slot in assignable_slots}
        room_usage_by_slot = {slot: set() for slot in assignable_slots}
        section_family_usage_by_slot: Dict[Tuple[int, int], Dict[str, str]] = {
            slot: {}
            for slot in assignable_slots
        }
        daily_subject_usage: Dict[Tuple[str, int, str], int] = {}
        teacher_daily_usage: Dict[Tuple[str, int], int] = {}
        lab_days_by_family: Dict[str, Set[int]] = {}
        candidate_preference_hits = 0

        shuffled_lectures = rng.sample(lecture_pool, len(lecture_pool))
        shuffled_lectures.sort(
            key=lambda lecture: (
                -int(lecture["duration"]),
                len(_resolve_candidate_rooms_for_lecture(lecture, rooms, section_home_rooms)),
                rng.random(),
            )
        )
        success = True

        preassigned_slots = list(normalized_locked_slots) + list(normalized_fixed_slots)
        if preassigned_slots:
            preassigned_slots.sort(
                key=lambda item: (
                    int(item["day_index"]),
                    int(item["start_period"]),
                    str(item["section_key"]),
                    str(item["subject"]),
                )
            )
            remaining_lectures = list(shuffled_lectures)

            for preassigned_slot in preassigned_slots:
                section_key = str(preassigned_slot["section_key"])
                subject_name = str(preassigned_slot["subject"])
                teacher_name = str(preassigned_slot["teacher"])
                day_index = int(preassigned_slot["day_index"])
                start_period = int(preassigned_slot["start_period"])
                duration = int(preassigned_slot["duration"])
                forced_room = str(preassigned_slot.get("room", ""))
                is_locked = bool(preassigned_slot.get("is_locked", False))

                match_index = -1
                for index, lecture in enumerate(remaining_lectures):
                    if (
                        str(lecture["section_key"]) == section_key
                        and str(lecture["subject"]) == subject_name
                        and str(lecture["teacher"]) == teacher_name
                        and int(lecture["duration"]) == duration
                    ):
                        match_index = index
                        break

                if match_index < 0:
                    success = False
                    break

                lecture = remaining_lectures.pop(match_index)
                section_family_key = str(lecture.get("section_family_key", section_key))
                lab_bundle_key = str(lecture.get("lab_bundle_key", "")).strip()
                family_slot_marker = lab_bundle_key or f"{section_key}::SINGLE"
                block_slots = [(day_index, start_period + offset) for offset in range(duration)]

                if any(schedule_by_section[section_key][slot] is not None for slot in block_slots):
                    success = False
                    break

                if any(teacher_name in teacher_usage_by_slot[slot] for slot in block_slots):
                    success = False
                    break

                allowed_family_markers = {None}
                if lab_bundle_key:
                    allowed_family_markers.add(lab_bundle_key)
                if any(
                    section_family_usage_by_slot.get(slot, {}).get(section_family_key) not in allowed_family_markers
                    for slot in block_slots
                ):
                    success = False
                    break

                teacher_day_key = (teacher_name, day_index)
                current_teacher_load = teacher_daily_usage.get(teacher_day_key, 0)
                subject_day_key = (section_key, day_index, subject_name)
                current_daily_subject_count = daily_subject_usage.get(subject_day_key, 0)
                preferred_hits = 0

                if not is_locked:
                    unavailable_slots = normalized_unavailability.get(teacher_name, set())
                    if any(slot in unavailable_slots for slot in block_slots):
                        success = False
                        break

                    preferred_hits, avoid_hits = preference_hit_counts(teacher_name, block_slots)
                    if avoid_hits > 0:
                        success = False
                        break

                    if would_exceed_max_consecutive(
                        teacher_name,
                        day_index,
                        block_slots,
                        teacher_usage_by_slot,
                    ):
                        success = False
                        break

                    if would_violate_faculty_rest_gap(
                        teacher_name,
                        block_slots,
                        teacher_usage_by_slot,
                    ):
                        success = False
                        break

                    if current_teacher_load + duration > teacher_daily_limit[teacher_name]:
                        success = False
                        break

                    subject_daily_limit = daily_subject_limits[(section_key, subject_name)]
                    if current_daily_subject_count + duration > subject_daily_limit:
                        success = False
                        break

                room_candidates = _resolve_candidate_rooms_for_lecture(
                    lecture,
                    rooms=rooms,
                    section_home_rooms=section_home_rooms,
                )
                if forced_room:
                    if forced_room not in room_candidates:
                        success = False
                        break
                    room_candidates = [forced_room]

                if not room_candidates:
                    success = False
                    break

                placed = False
                for room in rng.sample(room_candidates, len(room_candidates)):
                    if any(room in room_usage_by_slot[slot] for slot in block_slots):
                        continue

                    for offset, slot in enumerate(block_slots):
                        schedule_by_section[section_key][slot] = {
                            "class_name": lecture["class_name"],
                            "standard": lecture["standard"],
                            "section": lecture["section"],
                            "subject": lecture["subject"],
                            "teacher": lecture["teacher"],
                            "room": room,
                            "is_lab": bool(lecture.get("is_lab", False)),
                            "session_id": lecture["session_id"],
                            "session_length": duration,
                            "session_part": offset + 1,
                            "lab_bundle_key": lab_bundle_key,
                        }

                        teacher_usage_by_slot[slot].add(teacher_name)
                        room_usage_by_slot[slot].add(room)
                        section_family_usage_by_slot[slot][section_family_key] = family_slot_marker

                    daily_subject_usage[subject_day_key] = current_daily_subject_count + duration
                    teacher_daily_usage[teacher_day_key] = current_teacher_load + duration
                    if bool(lecture.get("is_lab", False)):
                        lab_days_by_family.setdefault(section_family_key, set()).add(day_index)
                    candidate_preference_hits += preferred_hits
                    placed = True
                    break

                if not placed:
                    success = False
                    break

            if not success:
                continue

            shuffled_lectures = remaining_lectures

        bundled_labs_by_key: Dict[str, List[Dict[str, object]]] = {}
        non_bundled_lectures: List[Dict[str, object]] = []
        for lecture in shuffled_lectures:
            bundle_key = str(lecture.get("lab_bundle_key", "")).strip()
            if bundle_key and bool(lecture.get("is_lab", False)):
                bundled_labs_by_key.setdefault(bundle_key, []).append(lecture)
            else:
                non_bundled_lectures.append(lecture)

        bundled_lab_items = list(bundled_labs_by_key.items())
        bundled_lab_items.sort(
            key=lambda item: (
                min(
                    len(_resolve_candidate_rooms_for_lecture(member, rooms, section_home_rooms))
                    for member in item[1]
                ),
                rng.random(),
            )
        )

        for bundle_key, bundle_members in bundled_lab_items:
            bundle_members = sorted(
                bundle_members,
                key=lambda lecture: (
                    int(lecture.get("lab_bundle_order") or 0),
                    str(lecture.get("section", "")),
                    str(lecture.get("subject", "")),
                    str(lecture.get("teacher", "")),
                ),
            )

            bundle_section_family_key = str(
                bundle_members[0].get("section_family_key", bundle_members[0].get("section_key", ""))
            )
            bundle_subject_count = max(int(member.get("lab_subject_count") or 0) for member in bundle_members)
            prefer_same_day_for_bundle = bundle_subject_count > int(lab_same_day_subject_threshold)
            family_lab_days = lab_days_by_family.get(bundle_section_family_key, set())

            if len(bundle_members) != 3:
                success = False
                break

            if len({str(member.get("teacher", "")) for member in bundle_members}) != 3:
                success = False
                break

            bundle_durations = {int(member.get("duration", 1)) for member in bundle_members}
            if len(bundle_durations) != 1:
                success = False
                break

            bundle_duration = bundle_durations.pop()
            candidate_starts = rng.sample(
                start_slots_by_duration[bundle_duration],
                len(start_slots_by_duration[bundle_duration]),
            )
            bundle_preference_cache: Dict[Tuple[int, int], Tuple[int, int]] = {}
            for slot in candidate_starts:
                slot_block = [(slot[0], slot[1] + offset) for offset in range(bundle_duration)]
                preferred_hits = 0
                avoid_hits = 0
                for member in bundle_members:
                    member_preferred, member_avoid = preference_hit_counts(
                        str(member["teacher"]),
                        slot_block,
                    )
                    preferred_hits += member_preferred
                    avoid_hits += member_avoid
                bundle_preference_cache[slot] = (preferred_hits, avoid_hits)

            candidate_starts.sort(
                key=lambda slot: (
                    bundle_preference_cache.get(slot, (0, 0))[1],
                    -bundle_preference_cache.get(slot, (0, 0))[0],
                    0
                    if not family_lab_days
                    else (
                        0
                        if (
                            (prefer_same_day_for_bundle and slot[0] in family_lab_days)
                            or (not prefer_same_day_for_bundle and slot[0] not in family_lab_days)
                        )
                        else 1
                    ),
                    slot[1]
                    if (
                        first_lunch_period_index_by_day.get(slot[0]) is not None
                        and slot[1] < int(first_lunch_period_index_by_day[slot[0]])
                    )
                    else 1000 + slot[1],
                    rng.random(),
                )
            )

            placed_bundle = False
            has_used_day_candidate = bool(family_lab_days) and any(
                slot_day in family_lab_days for slot_day, _ in candidate_starts
            )
            has_unused_day_candidate = bool(family_lab_days) and any(
                slot_day not in family_lab_days for slot_day, _ in candidate_starts
            )
            for day_index, start_period in candidate_starts:
                if (
                    family_lab_days
                    and not prefer_same_day_for_bundle
                    and has_unused_day_candidate
                    and day_index in family_lab_days
                ):
                    continue

                if (
                    family_lab_days
                    and prefer_same_day_for_bundle
                    and has_used_day_candidate
                    and day_index not in family_lab_days
                ):
                    continue

                block_slots = [(day_index, start_period + offset) for offset in range(bundle_duration)]
                member_room_options: List[List[str]] = []
                teacher_day_updates: List[Tuple[Tuple[str, int], int]] = []
                subject_day_updates: List[Tuple[Tuple[str, int, str], int]] = []
                bundle_preferred_hits = 0
                bundle_conflict = False

                for member in bundle_members:
                    section_key = str(member["section_key"])
                    section_family_key = str(member.get("section_family_key", section_key))
                    teacher_name = str(member["teacher"])
                    subject_name = str(member["subject"])
                    member_bundle_key = str(member.get("lab_bundle_key", "")).strip()

                    if any(schedule_by_section[section_key][slot] is not None for slot in block_slots):
                        bundle_conflict = True
                        break

                    if any(teacher_name in teacher_usage_by_slot[slot] for slot in block_slots):
                        bundle_conflict = True
                        break

                    allowed_family_markers = {None}
                    if member_bundle_key:
                        allowed_family_markers.add(member_bundle_key)
                    if any(
                        section_family_usage_by_slot.get(slot, {}).get(section_family_key) not in allowed_family_markers
                        for slot in block_slots
                    ):
                        bundle_conflict = True
                        break

                    unavailable_slots = normalized_unavailability.get(teacher_name, set())
                    if any(slot in unavailable_slots for slot in block_slots):
                        bundle_conflict = True
                        break

                    preferred_hits, avoid_hits = preference_hit_counts(teacher_name, block_slots)
                    if avoid_hits > 0:
                        bundle_conflict = True
                        break
                    bundle_preferred_hits += preferred_hits

                    if would_exceed_max_consecutive(
                        teacher_name,
                        day_index,
                        block_slots,
                        teacher_usage_by_slot,
                    ):
                        bundle_conflict = True
                        break

                    if would_violate_faculty_rest_gap(
                        teacher_name,
                        block_slots,
                        teacher_usage_by_slot,
                    ):
                        bundle_conflict = True
                        break

                    teacher_day_key = (teacher_name, day_index)
                    current_teacher_load = teacher_daily_usage.get(teacher_day_key, 0)
                    if current_teacher_load + bundle_duration > teacher_daily_limit[teacher_name]:
                        bundle_conflict = True
                        break
                    teacher_day_updates.append((teacher_day_key, current_teacher_load + bundle_duration))

                    subject_day_key = (section_key, day_index, subject_name)
                    current_subject_load = daily_subject_usage.get(subject_day_key, 0)
                    subject_daily_limit = daily_subject_limits[(section_key, subject_name)]
                    if current_subject_load + bundle_duration > subject_daily_limit:
                        bundle_conflict = True
                        break
                    subject_day_updates.append((subject_day_key, current_subject_load + bundle_duration))

                    room_candidates = _resolve_candidate_rooms_for_lecture(
                        member,
                        rooms=rooms,
                        section_home_rooms=section_home_rooms,
                    )
                    room_candidates = [
                        room_name
                        for room_name in room_candidates
                        if all(room_name not in room_usage_by_slot[slot] for slot in block_slots)
                    ]
                    if not room_candidates:
                        bundle_conflict = True
                        break

                    member_room_options.append(room_candidates)

                if bundle_conflict:
                    continue

                selected_rooms = choose_distinct_rooms_for_bundle(member_room_options)
                if not selected_rooms:
                    continue

                for member_index, member in enumerate(bundle_members):
                    section_key = str(member["section_key"])
                    section_family_key = str(member.get("section_family_key", section_key))
                    teacher_name = str(member["teacher"])
                    room_name = selected_rooms[member_index]
                    member_bundle_key = str(member.get("lab_bundle_key", "")).strip()
                    family_slot_marker = member_bundle_key or f"{section_key}::SINGLE"

                    for offset, slot in enumerate(block_slots):
                        schedule_by_section[section_key][slot] = {
                            "class_name": member["class_name"],
                            "standard": member["standard"],
                            "section": member["section"],
                            "subject": member["subject"],
                            "teacher": member["teacher"],
                            "room": room_name,
                            "is_lab": bool(member.get("is_lab", False)),
                            "session_id": member["session_id"],
                            "session_length": bundle_duration,
                            "session_part": offset + 1,
                            "lab_bundle_key": member_bundle_key,
                        }

                        teacher_usage_by_slot[slot].add(teacher_name)
                        room_usage_by_slot[slot].add(room_name)
                        section_family_usage_by_slot[slot][section_family_key] = family_slot_marker

                for teacher_day_key, new_load in teacher_day_updates:
                    teacher_daily_usage[teacher_day_key] = new_load

                for subject_day_key, new_count in subject_day_updates:
                    daily_subject_usage[subject_day_key] = new_count

                lab_days_by_family.setdefault(bundle_section_family_key, set()).add(day_index)

                candidate_preference_hits += bundle_preferred_hits
                placed_bundle = True
                break

            if not placed_bundle:
                success = False
                break

        if not success:
            continue

        shuffled_lectures = non_bundled_lectures

        for lecture in shuffled_lectures:
            section_key = str(lecture["section_key"])
            section_family_key = str(lecture.get("section_family_key", section_key))
            teacher_name = str(lecture["teacher"])
            subject_name = str(lecture["subject"])
            lab_bundle_key = str(lecture.get("lab_bundle_key", "")).strip()
            family_slot_marker = lab_bundle_key or f"{section_key}::SINGLE"
            duration = int(lecture["duration"])
            candidate_starts = rng.sample(start_slots_by_duration[duration], len(start_slots_by_duration[duration]))
            preference_cache: Dict[Tuple[int, int], Tuple[int, int]] = {
                slot: preference_hit_counts(
                    teacher_name,
                    [(slot[0], slot[1] + offset) for offset in range(duration)],
                )
                for slot in candidate_starts
            }
            candidate_starts.sort(
                key=lambda slot: (
                    preference_cache.get(slot, (0, 0))[1],
                    -preference_cache.get(slot, (0, 0))[0],
                    slot[1]  # Strongly prefer earlier periods (earlier = earlier in pre-lunch)
                    if (
                        first_lunch_period_index_by_day.get(slot[0]) is not None
                        and slot[1] < int(first_lunch_period_index_by_day[slot[0]])
                    )
                    else 1000 + slot[1],  # Post-lunch slots get much higher priority value
                    rng.random(),
                )
            )
            placed = False

            for day_index, start_period in candidate_starts:
                block_slots = [(day_index, start_period + offset) for offset in range(duration)]

                if any(schedule_by_section[section_key][slot] is not None for slot in block_slots):
                    continue

                if any(teacher_name in teacher_usage_by_slot[slot] for slot in block_slots):
                    continue

                allowed_family_markers = {None}
                if lab_bundle_key:
                    allowed_family_markers.add(lab_bundle_key)
                if any(
                    section_family_usage_by_slot.get(slot, {}).get(section_family_key) not in allowed_family_markers
                    for slot in block_slots
                ):
                    continue

                unavailable_slots = normalized_unavailability.get(teacher_name, set())
                if any(slot in unavailable_slots for slot in block_slots):
                    continue

                preferred_hits, avoid_hits = preference_cache.get((day_index, start_period), (0, 0))
                if avoid_hits > 0:
                    continue

                if would_exceed_max_consecutive(
                    teacher_name,
                    day_index,
                    block_slots,
                    teacher_usage_by_slot,
                ):
                    continue

                if would_violate_faculty_rest_gap(
                    teacher_name,
                    block_slots,
                    teacher_usage_by_slot,
                ):
                    continue

                teacher_day_key = (teacher_name, day_index)
                current_teacher_load = teacher_daily_usage.get(teacher_day_key, 0)
                if current_teacher_load + duration > teacher_daily_limit[teacher_name]:
                    continue

                subject_day_key = (section_key, day_index, subject_name)
                current_daily_subject_count = daily_subject_usage.get(subject_day_key, 0)
                subject_daily_limit = daily_subject_limits[(section_key, subject_name)]
                if current_daily_subject_count + duration > subject_daily_limit:
                    continue

                room_candidates = _resolve_candidate_rooms_for_lecture(
                    lecture,
                    rooms=rooms,
                    section_home_rooms=section_home_rooms,
                )
                if not room_candidates:
                    continue

                for room in rng.sample(room_candidates, len(room_candidates)):
                    if any(room in room_usage_by_slot[slot] for slot in block_slots):
                        continue

                    for offset, slot in enumerate(block_slots):
                        schedule_by_section[section_key][slot] = {
                            "class_name": lecture["class_name"],
                            "standard": lecture["standard"],
                            "section": lecture["section"],
                            "subject": lecture["subject"],
                            "teacher": lecture["teacher"],
                            "room": room,
                            "is_lab": bool(lecture.get("is_lab", False)),
                            "session_id": lecture["session_id"],
                            "session_length": duration,
                            "session_part": offset + 1,
                            "lab_bundle_key": lab_bundle_key,
                        }

                        teacher_usage_by_slot[slot].add(teacher_name)
                        room_usage_by_slot[slot].add(room)
                        section_family_usage_by_slot[slot][section_family_key] = family_slot_marker

                    daily_subject_usage[subject_day_key] = current_daily_subject_count + duration
                    teacher_daily_usage[teacher_day_key] = current_teacher_load + duration
                    candidate_preference_hits += preferred_hits
                    placed = True
                    break

                if placed:
                    break

            if not placed:
                success = False
                break

        if not success:
            continue

        section_tables_candidate = _build_section_tables_from_schedule(
            schedule_by_section,
            section_metadata,
            working_days,
            periods_per_day,
        )
        score, penalties = _calculate_soft_score(
            section_tables_candidate,
            working_days,
            periods_per_day,
            lunch_by_day,
        )
        score += min(3.0, candidate_preference_hits * 0.03)
        successful_candidates += 1

        if score > best_score:
            best_score = score
            best_penalties = penalties
            best_section_tables = section_tables_candidate
            stagnant_successes = 0

            if score >= 99.5:
                break
        else:
            stagnant_successes += 1

        if (
            successful_candidates >= target_successful_candidates
            and stagnant_successes >= stagnation_limit
        ):
            break

    if best_section_tables is None:
        raise ValueError(
            "Unable to generate a conflict-free timetable for all sections after multiple retries. "
            "Try adding rooms, increasing days/periods, or reducing lecture load."
        )

    optimization = {
        "score": round(best_score, 2),
        "penalties": best_penalties,
        "preference_faculty_count": len(normalized_preferences),
        "locked_slot_blocks": len(normalized_locked_slots),
        "require_daily_free_period": bool(require_daily_free_period),
        "successful_candidates": successful_candidates,
        "attempts": attempts_used,
        "attempts_budget": effective_retry_budget,
        "strategy": "best-of-randomized-candidates",
    }

    return day_labels, period_labels, best_section_tables, optimization


def multi_timetable_to_long_dataframe(
    days: List[str],
    periods: List[str],
    section_tables: List[SectionTable],
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> pd.DataFrame:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    rows: List[Dict[str, object]] = []

    for table in section_tables:
        matrix = table["timetable"]
        for day_index, day_name in enumerate(days):
            for period_index, period_name in enumerate(periods):
                period_number = period_index + 1
                cell = matrix[day_index][period_index]

                if _is_lunch_period(lunch_by_day, day_index, period_number):
                    rows.append(
                        {
                            "Class": table["class_name"],
                            "Standard": table["standard"],
                            "Section": table["section"],
                            "SectionKey": table["section_key"],
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "LUNCH BREAK",
                            "Teacher": "",
                            "Room": "",
                            "IsLab": False,
                            "IsLunch": True,
                            "IsFree": False,
                            "LectureType": "LUNCH",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                if cell is None:
                    rows.append(
                        {
                            "Class": table["class_name"],
                            "Standard": table["standard"],
                            "Section": table["section"],
                            "SectionKey": table["section_key"],
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "",
                            "Teacher": "",
                            "Room": "",
                            "IsLab": False,
                            "IsLunch": False,
                            "IsFree": True,
                            "LectureType": "FREE",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                is_lab = bool(cell.get("is_lab", False))
                rows.append(
                    {
                        "Class": table["class_name"],
                        "Standard": table["standard"],
                        "Section": table["section"],
                        "SectionKey": table["section_key"],
                        "Day": day_name,
                        "Period": period_name,
                        "Subject": str(cell.get("subject", "")),
                        "Teacher": str(cell.get("teacher", "")),
                        "Room": str(cell.get("room", "")),
                        "IsLab": is_lab,
                        "IsLunch": False,
                        "IsFree": False,
                        "LectureType": "LAB" if is_lab else "THEORY",
                        "SessionLength": int(cell.get("session_length", 1)),
                        "SessionPart": int(cell.get("session_part", 1)),
                    }
                )

    return pd.DataFrame(rows)


def faculty_timetable_to_long_dataframe(
    days: List[str],
    periods: List[str],
    faculty_tables: List[FacultyTable],
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> pd.DataFrame:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    rows: List[Dict[str, object]] = []

    for table in faculty_tables:
        matrix: TimetableMatrix = table["timetable"]
        teacher_name = str(table["teacher"])

        for day_index, day_name in enumerate(days):
            for period_index, period_name in enumerate(periods):
                period_number = period_index + 1
                cell = matrix[day_index][period_index]

                if _is_lunch_period(lunch_by_day, day_index, period_number):
                    rows.append(
                        {
                            "Teacher": teacher_name,
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "LUNCH BREAK",
                            "Class": "",
                            "Standard": "",
                            "Section": "",
                            "SectionKey": "",
                            "Room": "",
                            "IsLab": False,
                            "IsLunch": True,
                            "IsFree": False,
                            "HasConflict": False,
                            "LectureType": "LUNCH",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                if cell is None:
                    rows.append(
                        {
                            "Teacher": teacher_name,
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "",
                            "Class": "",
                            "Standard": "",
                            "Section": "",
                            "SectionKey": "",
                            "Room": "",
                            "IsLab": False,
                            "IsLunch": False,
                            "IsFree": True,
                            "HasConflict": False,
                            "LectureType": "FREE",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                if isinstance(cell, dict) and "conflicts" in cell:
                    conflict_items = list(cell.get("conflicts", []))
                    rows.append(
                        {
                            "Teacher": teacher_name,
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "CONFLICT",
                            "Class": "",
                            "Standard": "",
                            "Section": "",
                            "SectionKey": "",
                            "Room": "",
                            "IsLab": any(bool(item.get("is_lab", False)) for item in conflict_items),
                            "IsLunch": False,
                            "IsFree": False,
                            "HasConflict": True,
                            "LectureType": "CONFLICT",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                rows.append(
                    {
                        "Teacher": teacher_name,
                        "Day": day_name,
                        "Period": period_name,
                        "Subject": str(cell.get("subject", "")),
                        "Class": str(cell.get("class_name", "")),
                        "Standard": str(cell.get("standard", "")),
                        "Section": str(cell.get("section", "")),
                        "SectionKey": str(cell.get("section_key", "")),
                        "Room": str(cell.get("room", "")),
                        "IsLab": bool(cell.get("is_lab", False)),
                        "IsLunch": False,
                        "IsFree": False,
                        "HasConflict": False,
                        "LectureType": "LAB" if bool(cell.get("is_lab", False)) else "THEORY",
                        "SessionLength": int(cell.get("session_length", 1)),
                        "SessionPart": int(cell.get("session_part", 1)),
                    }
                )

    return pd.DataFrame(rows)


def class_timetable_to_long_dataframe(
    days: List[str],
    periods: List[str],
    class_tables: List[ClassTable],
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> pd.DataFrame:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    rows: List[Dict[str, object]] = []

    for class_table in class_tables:
        class_name = str(class_table["class_name"])
        standard = str(class_table["standard"])
        class_key = str(class_table["class_key"])

        for section_table in class_table["sections"]:
            section_name = str(section_table["section"])
            section_key = str(section_table["section_key"])
            matrix: TimetableMatrix = section_table["timetable"]

            for day_index, day_name in enumerate(days):
                for period_index, period_name in enumerate(periods):
                    period_number = period_index + 1
                    cell = matrix[day_index][period_index]

                    if _is_lunch_period(lunch_by_day, day_index, period_number):
                        rows.append(
                            {
                                "Class": class_name,
                                "Standard": standard,
                                "ClassKey": class_key,
                                "Section": section_name,
                                "SectionKey": section_key,
                                "Day": day_name,
                                "Period": period_name,
                                "Subject": "LUNCH BREAK",
                                "Teacher": "",
                                "Room": "",
                                "IsLab": False,
                                "IsLunch": True,
                                "IsFree": False,
                                "LectureType": "LUNCH",
                                "SessionLength": 0,
                                "SessionPart": 0,
                            }
                        )
                        continue

                    if cell is None:
                        rows.append(
                            {
                                "Class": class_name,
                                "Standard": standard,
                                "ClassKey": class_key,
                                "Section": section_name,
                                "SectionKey": section_key,
                                "Day": day_name,
                                "Period": period_name,
                                "Subject": "",
                                "Teacher": "",
                                "Room": "",
                                "IsLab": False,
                                "IsLunch": False,
                                "IsFree": True,
                                "LectureType": "FREE",
                                "SessionLength": 0,
                                "SessionPart": 0,
                            }
                        )
                        continue

                    is_lab = bool(cell.get("is_lab", False))
                    rows.append(
                        {
                            "Class": class_name,
                            "Standard": standard,
                            "ClassKey": class_key,
                            "Section": section_name,
                            "SectionKey": section_key,
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": str(cell.get("subject", "")),
                            "Teacher": str(cell.get("teacher", "")),
                            "Room": str(cell.get("room", "")),
                            "IsLab": is_lab,
                            "IsLunch": False,
                            "IsFree": False,
                            "LectureType": "LAB" if is_lab else "THEORY",
                            "SessionLength": int(cell.get("session_length", 1)),
                            "SessionPart": int(cell.get("session_part", 1)),
                        }
                    )

    return pd.DataFrame(rows)


def room_timetable_to_long_dataframe(
    days: List[str],
    periods: List[str],
    room_tables: List[RoomTable],
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> pd.DataFrame:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    rows: List[Dict[str, object]] = []

    for table in room_tables:
        matrix: TimetableMatrix = table["timetable"]
        room_name = str(table["room"])

        for day_index, day_name in enumerate(days):
            for period_index, period_name in enumerate(periods):
                period_number = period_index + 1
                cell = matrix[day_index][period_index]

                if _is_lunch_period(lunch_by_day, day_index, period_number):
                    rows.append(
                        {
                            "Room": room_name,
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "LUNCH BREAK",
                            "Teacher": "",
                            "Class": "",
                            "Standard": "",
                            "Section": "",
                            "SectionKey": "",
                            "IsLab": False,
                            "IsLunch": True,
                            "IsFree": False,
                            "HasConflict": False,
                            "LectureType": "LUNCH",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                if cell is None:
                    rows.append(
                        {
                            "Room": room_name,
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "",
                            "Teacher": "",
                            "Class": "",
                            "Standard": "",
                            "Section": "",
                            "SectionKey": "",
                            "IsLab": False,
                            "IsLunch": False,
                            "IsFree": True,
                            "HasConflict": False,
                            "LectureType": "FREE",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                if isinstance(cell, dict) and "conflicts" in cell:
                    conflict_items = list(cell.get("conflicts", []))
                    rows.append(
                        {
                            "Room": room_name,
                            "Day": day_name,
                            "Period": period_name,
                            "Subject": "CONFLICT",
                            "Teacher": "",
                            "Class": "",
                            "Standard": "",
                            "Section": "",
                            "SectionKey": "",
                            "IsLab": any(bool(item.get("is_lab", False)) for item in conflict_items),
                            "IsLunch": False,
                            "IsFree": False,
                            "HasConflict": True,
                            "LectureType": "CONFLICT",
                            "SessionLength": 0,
                            "SessionPart": 0,
                        }
                    )
                    continue

                rows.append(
                    {
                        "Room": room_name,
                        "Day": day_name,
                        "Period": period_name,
                        "Subject": str(cell.get("subject", "")),
                        "Teacher": str(cell.get("teacher", "")),
                        "Class": str(cell.get("class_name", "")),
                        "Standard": str(cell.get("standard", "")),
                        "Section": str(cell.get("section", "")),
                        "SectionKey": str(cell.get("section_key", "")),
                        "IsLab": bool(cell.get("is_lab", False)),
                        "IsLunch": False,
                        "IsFree": False,
                        "HasConflict": False,
                        "LectureType": "LAB" if bool(cell.get("is_lab", False)) else "THEORY",
                        "SessionLength": int(cell.get("session_length", 1)),
                        "SessionPart": int(cell.get("session_part", 1)),
                    }
                )

    return pd.DataFrame(rows)


def build_multi_timetable_summary(
    days: List[str],
    periods: List[str],
    section_tables: List[SectionTable],
    room_names: List[str] | None = None,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
    optimization: Dict[str, object] | None = None,
) -> Dict[str, object]:
    day_count = len(days)
    period_count = len(periods)
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        day_count,
        period_count,
    )
    lunch_set = _normalize_lunch_break_periods(
        lunch_break_periods,
        period_count,
        day_count,
    )

    subject_counter: Counter[str] = Counter()
    teacher_counter: Counter[str] = Counter()
    room_counter: Counter[str] = Counter()
    day_counter: Counter[str] = Counter()
    section_loads: Dict[str, int] = {}
    section_stats: Dict[str, Dict[str, object]] = {}
    class_rollups: Dict[str, Dict[str, object]] = {}

    total_slots_per_section = day_count * period_count
    lunch_slots_per_section = sum(
        len(lunch_by_day.get(day_index, set()))
        for day_index in range(day_count)
    )
    teaching_slots_per_section = total_slots_per_section - lunch_slots_per_section

    total_slots = teaching_slots_per_section * len(section_tables)
    scheduled_slots = 0

    for table in section_tables:
        section_key = str(table["section_key"])
        matrix: TimetableMatrix = table["timetable"]
        section_scheduled = 0

        for day_index, day_name in enumerate(days):
            for period_index in range(len(periods)):
                period_number = period_index + 1
                if _is_lunch_period(lunch_by_day, day_index, period_number):
                    continue

                cell = matrix[day_index][period_index]
                if cell is None:
                    continue

                section_scheduled += 1
                scheduled_slots += 1
                subject_counter[str(cell["subject"])] += 1
                teacher_counter[str(cell["teacher"])] += 1
                room_counter[str(cell["room"])] += 1
                day_counter[day_name] += 1

        section_loads[section_key] = section_scheduled
        section_stats[section_key] = {
            "class_name": table["class_name"],
            "standard": table["standard"],
            "section": table["section"],
            "scheduled_slots": section_scheduled,
            "total_slots": teaching_slots_per_section,
            "lunch_slots": lunch_slots_per_section,
            "free_slots": teaching_slots_per_section - section_scheduled,
            "utilization_percent": round(
                (section_scheduled / teaching_slots_per_section) * 100, 2
            ) if teaching_slots_per_section else 0.0,
        }

        class_key = build_class_key(str(table["class_name"]), str(table["standard"]))
        if class_key not in class_rollups:
            class_rollups[class_key] = {
                "class_name": str(table["class_name"]),
                "standard": str(table["standard"]),
                "section_count": 0,
                "scheduled_slots": 0,
                "total_slots": 0,
            }

        class_rollups[class_key]["section_count"] = int(class_rollups[class_key]["section_count"]) + 1
        class_rollups[class_key]["scheduled_slots"] = int(class_rollups[class_key]["scheduled_slots"]) + section_scheduled
        class_rollups[class_key]["total_slots"] = int(class_rollups[class_key]["total_slots"]) + teaching_slots_per_section

    free_slots = total_slots - scheduled_slots
    utilization_percent = round((scheduled_slots / total_slots) * 100, 2) if total_slots else 0.0
    effective_room_names = sorted({str(room) for room in (room_names or list(room_counter.keys())) if str(room).strip()})
    room_total_slots = teaching_slots_per_section
    room_stats = {
        room_name: {
            "scheduled_slots": room_counter.get(room_name, 0),
            "total_slots": room_total_slots,
            "free_slots": max(0, room_total_slots - room_counter.get(room_name, 0)),
            "utilization_percent": round((room_counter.get(room_name, 0) / room_total_slots) * 100, 2) if room_total_slots else 0.0,
        }
        for room_name in effective_room_names
    }
    class_stats = {
        class_key: {
            "class_name": str(class_rollups[class_key]["class_name"]),
            "standard": str(class_rollups[class_key]["standard"]),
            "section_count": int(class_rollups[class_key]["section_count"]),
            "scheduled_slots": int(class_rollups[class_key]["scheduled_slots"]),
            "total_slots": int(class_rollups[class_key]["total_slots"]),
            "free_slots": int(class_rollups[class_key]["total_slots"]) - int(class_rollups[class_key]["scheduled_slots"]),
            "utilization_percent": round(
                (int(class_rollups[class_key]["scheduled_slots"]) / int(class_rollups[class_key]["total_slots"])) * 100,
                2,
            ) if int(class_rollups[class_key]["total_slots"]) else 0.0,
        }
        for class_key in sorted(class_rollups.keys())
    }
    faculty_total_slots = teaching_slots_per_section
    faculty_stats = {
        teacher: {
            "scheduled_slots": load,
            "total_slots": faculty_total_slots,
            "free_slots": max(0, faculty_total_slots - load),
            "utilization_percent": round((load / faculty_total_slots) * 100, 2) if faculty_total_slots else 0.0,
        }
        for teacher, load in sorted(teacher_counter.items())
    }

    summary = {
        "section_count": len(section_tables),
        "class_count": len(class_stats),
        "faculty_count": len(teacher_counter),
        "room_count": len(effective_room_names),
        "total_slots": total_slots,
        "teaching_slots_per_section": teaching_slots_per_section,
        "lunch_slots": lunch_slots_per_section * len(section_tables),
        "lunch_break_periods": sorted(lunch_set),
        "lunch_break_periods_by_day": {
            str(day_index + 1): sorted(lunch_by_day.get(day_index, set()))
            for day_index in range(day_count)
        },
        "scheduled_slots": scheduled_slots,
        "free_slots": free_slots,
        "utilization_percent": utilization_percent,
        "subject_loads": dict(subject_counter.most_common()),
        "teacher_loads": dict(teacher_counter.most_common()),
        "room_loads": dict(room_counter.most_common()),
        "daily_loads": {day: day_counter.get(day, 0) for day in days},
        "section_loads": section_loads,
        "section_stats": section_stats,
        "class_stats": class_stats,
        "faculty_stats": faculty_stats,
        "room_stats": room_stats,
    }

    if optimization is not None:
        summary["optimization"] = optimization

    return summary


def validate_multi_section_timetable(
    section_tables: List[SectionTable],
    assignments: List[FacultyAssignment],
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
    faculty_daily_default_max: int | None = None,
    faculty_daily_limits: Dict[str, int] | None = None,
    faculty_unavailability: Dict[str, Set[Tuple[int, int]]] | None = None,
    faculty_preferences: Dict[str, Dict[str, object]] | None = None,
    fixed_slot_constraints: List[Dict[str, object]] | None = None,
    require_daily_free_period: bool = False,
) -> List[str]:
    issues: List[str] = []

    if not section_tables:
        return ["No section timetable generated."]

    try:
        assignments = _expand_assignments_for_lab_groups(assignments)
    except ValueError as error:
        return [str(error)]

    day_count = len(section_tables[0]["timetable"])
    period_count = len(section_tables[0]["timetable"][0]) if day_count else 0
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        day_count,
        period_count,
    )
    faculty_unavailability = faculty_unavailability or {}
    faculty_preferences = faculty_preferences or {}
    fixed_slot_constraints = fixed_slot_constraints or []

    normalized_unavailability: Dict[str, Set[Tuple[int, int]]] = {}
    for teacher_name, slot_items in faculty_unavailability.items():
        teacher_key = str(teacher_name).strip()
        if not teacher_key:
            continue

        teacher_slots: Set[Tuple[int, int]] = set()
        for raw_slot in slot_items:
            if not isinstance(raw_slot, tuple) or len(raw_slot) != 2:
                issues.append(
                    f"Invalid unavailable slot configured for faculty '{teacher_key}'."
                )
                continue

            try:
                day_index = int(raw_slot[0])
                period_index = int(raw_slot[1])
            except (TypeError, ValueError):
                issues.append(
                    f"Invalid unavailable slot configured for faculty '{teacher_key}'."
                )
                continue

            if day_index < 0 or day_index >= day_count or period_index < 0 or period_index >= period_count:
                issues.append(
                    f"Unavailable slot for faculty '{teacher_key}' is out of timetable range: "
                    f"day {day_index + 1}, period {period_index + 1}."
                )
                continue

            teacher_slots.add((day_index, period_index))

        if teacher_slots:
            normalized_unavailability[teacher_key] = teacher_slots

    normalized_preferences: Dict[str, Dict[str, object]] = {}
    for teacher_name, raw_config in faculty_preferences.items():
        teacher_key = str(teacher_name).strip()
        if not teacher_key:
            continue

        config = raw_config if isinstance(raw_config, dict) else {}
        preferred_slots = {
            (int(slot[0]), int(slot[1]))
            for slot in config.get("preferred_slots", set())
            if isinstance(slot, tuple) and len(slot) == 2
        }
        avoid_slots = {
            (int(slot[0]), int(slot[1]))
            for slot in config.get("avoid_slots", set())
            if isinstance(slot, tuple) and len(slot) == 2
        }

        max_consecutive = config.get("max_consecutive")
        normalized_preferences[teacher_key] = {
            "preferred_slots": preferred_slots,
            "avoid_slots": avoid_slots,
            "max_consecutive": int(max_consecutive) if max_consecutive is not None else None,
        }

    normalized_fixed_slots: List[Dict[str, object]] = []
    for index, raw_constraint in enumerate(fixed_slot_constraints, start=1):
        try:
            normalized_fixed_slots.append(
                {
                    "section_key": str(raw_constraint["section_key"]),
                    "subject": str(raw_constraint["subject"]),
                    "teacher": str(raw_constraint["teacher"]),
                    "day_index": int(raw_constraint["day_index"]),
                    "start_period": int(raw_constraint["start_period"]),
                    "duration": int(raw_constraint.get("duration", 1) or 1),
                    "room": str(raw_constraint.get("room", "")).strip(),
                }
            )
        except (KeyError, TypeError, ValueError):
            issues.append(f"Fixed-slot constraint #{index} is invalid.")

    table_by_section = {str(table["section_key"]): table for table in section_tables}

    expected_counts = _expected_subject_counts(assignments)
    actual_counts: Dict[Tuple[str, str], int] = {}

    faculty_daily_limits = faculty_daily_limits or {}
    max_teaching_periods_per_day = max(
        1,
        min(
            period_count - len(lunch_by_day.get(day_index, set()))
            for day_index in range(max(1, day_count))
        ),
    )
    default_limit = (
        max_teaching_periods_per_day
        if faculty_daily_default_max is None
        else max(1, min(faculty_daily_default_max, max_teaching_periods_per_day))
    )

    section_subject_daily_override: Dict[Tuple[str, str], int] = {}
    section_subject_is_lab: Dict[Tuple[str, str], bool] = {}
    for assignment in assignments:
        section_key = build_section_key(
            str(assignment["class_name"]),
            str(assignment["standard"]),
            str(assignment["section"]),
        )
        subject_key = (section_key, str(assignment["subject"]))
        section_subject_is_lab[subject_key] = section_subject_is_lab.get(subject_key, False) or bool(
            assignment.get("is_lab", False)
        )
        override = assignment.get("subject_daily_max")
        if override is not None:
            section_subject_daily_override[subject_key] = int(override)

    section_subject_expected = _expected_subject_counts(assignments)
    section_subject_daily_limits: Dict[Tuple[str, str], int] = {}
    for key, expected_periods in section_subject_expected.items():
        computed_default = max(1, math.ceil(expected_periods / max(1, day_count)))
        if section_subject_is_lab.get(key, False):
            computed_default = max(2, computed_default)

        chosen = section_subject_daily_override.get(key, computed_default)
        if section_subject_is_lab.get(key, False):
            section_subject_daily_limits[key] = max(2, min(chosen, max_teaching_periods_per_day))
        else:
            section_subject_daily_limits[key] = max(1, min(chosen, max_teaching_periods_per_day))

    teacher_daily_counter: Dict[Tuple[str, int], int] = {}
    teacher_periods_by_day: Dict[Tuple[str, int], Set[int]] = {}
    teacher_cells_by_day: Dict[Tuple[str, int], Dict[int, TimetableCell]] = {}
    teacher_limit_map: Dict[str, int] = {}
    teacher_has_theory_load: Dict[str, bool] = {}
    section_subject_day_counter: Dict[Tuple[str, int, str], int] = {}
    slot_family_cells: Dict[Tuple[int, int], Dict[str, List[Tuple[str, TimetableCell]]]] = {}
    lab_session_occurrences: Dict[Tuple[str, str], List[Tuple[int, int, TimetableCell]]] = {}
    lab_bundle_occurrences: Dict[str, List[Tuple[str, int, int, TimetableCell]]] = {}

    for assignment in assignments:
        teacher_name = str(assignment["teacher"])
        if not bool(assignment.get("is_lab", False)):
            teacher_has_theory_load[teacher_name] = True
        else:
            teacher_has_theory_load.setdefault(teacher_name, False)

        custom_limit = faculty_daily_limits.get(teacher_name, default_limit)
        effective_limit = min(max(1, custom_limit), max_teaching_periods_per_day)
        if (
            require_daily_free_period
            and max_teaching_periods_per_day > 1
            and teacher_has_theory_load.get(teacher_name, False)
        ):
            effective_limit = min(effective_limit, max_teaching_periods_per_day - 1)
        teacher_limit_map[teacher_name] = max(1, effective_limit)

    for day_index in range(day_count):
        for period_index in range(period_count):
            period_number = period_index + 1

            if _is_lunch_period(lunch_by_day, day_index, period_number):
                for table in section_tables:
                    matrix: TimetableMatrix = table["timetable"]
                    if matrix[day_index][period_index] is not None:
                        issues.append(
                            f"Lecture found during lunch break in {table['section_key']} on day {day_index + 1}, period {period_number}."
                        )
                continue

            teacher_set = set()
            room_set = set()

            for table in section_tables:
                matrix = table["timetable"]
                cell = matrix[day_index][period_index]
                if cell is None:
                    continue

                for required_key in (
                    "class_name",
                    "standard",
                    "section",
                    "subject",
                    "teacher",
                    "room",
                ):
                    if required_key not in cell or not str(cell[required_key]).strip():
                        issues.append(
                            f"Missing {required_key} at {table['section_key']} on day {day_index + 1}, period {period_number}."
                        )

                teacher = str(cell.get("teacher", ""))
                if teacher in teacher_set:
                    issues.append(
                        f"Teacher conflict for {teacher} on day {day_index + 1}, period {period_number}."
                    )
                teacher_set.add(teacher)

                room = str(cell.get("room", ""))
                if room in room_set:
                    issues.append(
                        f"Room conflict for {room} on day {day_index + 1}, period {period_number}."
                    )
                room_set.add(room)

                subject_name = str(cell.get("subject", ""))
                section_key = str(table["section_key"])

                section_subject_key = (section_key, subject_name)
                actual_counts[section_subject_key] = actual_counts.get(section_subject_key, 0) + 1

                subject_day_key = (section_key, day_index, subject_name)
                section_subject_day_counter[subject_day_key] = section_subject_day_counter.get(subject_day_key, 0) + 1

                teacher_day_key = (teacher, day_index)
                teacher_daily_counter[teacher_day_key] = teacher_daily_counter.get(teacher_day_key, 0) + 1
                teacher_periods_by_day.setdefault(teacher_day_key, set()).add(period_index)
                teacher_cells_by_day.setdefault(teacher_day_key, {})[period_index] = cell

                section_family_key = _build_section_family_key(
                    str(cell.get("class_name", table.get("class_name", ""))),
                    str(cell.get("standard", table.get("standard", ""))),
                    str(cell.get("section", table.get("section", ""))),
                )
                slot_family_cells.setdefault((day_index, period_index), {}).setdefault(
                    section_family_key,
                    [],
                ).append((section_key, cell))

                unavailable_slots = normalized_unavailability.get(teacher, set())
                if (day_index, period_index) in unavailable_slots:
                    issues.append(
                        f"Faculty {teacher} assigned during unavailable slot on day {day_index + 1}, period {period_number}."
                    )

                avoid_slots = set(normalized_preferences.get(teacher, {}).get("avoid_slots", set()))
                if (day_index, period_index) in avoid_slots:
                    issues.append(
                        f"Faculty {teacher} assigned in avoided slot on day {day_index + 1}, period {period_number}."
                    )

                if bool(cell.get("is_lab", False)):
                    session_id = str(cell.get("session_id", ""))
                    if not session_id:
                        issues.append(
                            f"Lab session missing session_id at {section_key} day {day_index + 1}, period {period_number}."
                        )
                        continue
                    lab_key = (section_key, session_id)
                    lab_session_occurrences.setdefault(lab_key, []).append((day_index, period_index, cell))

                    bundle_key = str(cell.get("lab_bundle_key", "")).strip()
                    if bundle_key:
                        lab_bundle_occurrences.setdefault(bundle_key, []).append(
                            (section_key, day_index, period_index, cell)
                        )

    for section_subject_key, expected_value in expected_counts.items():
        actual_value = actual_counts.get(section_subject_key, 0)
        if actual_value != expected_value:
            issues.append(
                f"{section_subject_key[0]} -> {section_subject_key[1]} has {actual_value} lectures but expected {expected_value}."
            )

    for (section_key, day_index, subject_name), count in section_subject_day_counter.items():
        max_allowed = section_subject_daily_limits.get((section_key, subject_name), max_teaching_periods_per_day)
        if count > max_allowed:
            issues.append(
                f"{section_key} exceeds daily max for {subject_name} on day {day_index + 1}: {count}>{max_allowed}."
            )

    for (teacher_name, day_index), actual_load in teacher_daily_counter.items():
        max_limit = teacher_limit_map.get(teacher_name, default_limit)
        if actual_load > max_limit:
            issues.append(
                f"Faculty {teacher_name} exceeds daily max periods on day {day_index + 1}: {actual_load}>{max_limit}."
            )

        if require_daily_free_period and teacher_has_theory_load.get(teacher_name, False):
            day_teaching_capacity = period_count - len(lunch_by_day.get(day_index, set()))
            if day_teaching_capacity > 1 and actual_load >= day_teaching_capacity:
                issues.append(
                    f"Faculty {teacher_name} has no free period on day {day_index + 1}."
                )

    for (day_index, period_index), family_entries in slot_family_cells.items():
        for family_key, entries in family_entries.items():
            if len(entries) <= 1:
                continue

            family_cells = [cell for _, cell in entries]
            bundle_keys = {str(cell.get("lab_bundle_key", "")).strip() for cell in family_cells}
            group_numbers: Set[int] = set()
            has_only_group_sections = True
            for _, cell in entries:
                _, group_number = _extract_section_group(str(cell.get("section", "")))
                if group_number is None:
                    has_only_group_sections = False
                    break
                group_numbers.add(group_number)

            is_valid_group_lab_overlap = (
                len(entries) == 3
                and all(bool(cell.get("is_lab", False)) for cell in family_cells)
                and len(bundle_keys) == 1
                and "" not in bundle_keys
                and has_only_group_sections
                and group_numbers == {1, 2, 3}
            )
            if not is_valid_group_lab_overlap:
                issues.append(
                    f"Section family conflict for {family_key} on day {day_index + 1}, period {period_index + 1}."
                )

    for (teacher_name, day_index), period_cells in teacher_cells_by_day.items():
        sorted_periods = sorted(period_cells.keys())
        for index in range(1, len(sorted_periods)):
            previous_period = sorted_periods[index - 1]
            current_period = sorted_periods[index]
            if current_period != previous_period + 1:
                continue

            first_cell = period_cells[previous_period]
            second_cell = period_cells[current_period]
            if _cells_form_single_lab_block(first_cell, second_cell):
                continue

            issues.append(
                f"Faculty {teacher_name} has consecutive periods on day {day_index + 1}: "
                f"{previous_period + 1}-{current_period + 1}."
            )

    for (teacher_name, day_index), period_set in teacher_periods_by_day.items():
        preference_config = normalized_preferences.get(teacher_name, {})
        max_consecutive = preference_config.get("max_consecutive")
        if max_consecutive is None:
            continue

        run_length = 0
        previous_period = None
        for period_index in sorted(period_set):
            if previous_period is None or period_index == previous_period + 1:
                run_length += 1
            else:
                run_length = 1

            if run_length > int(max_consecutive):
                issues.append(
                    f"Faculty {teacher_name} exceeds max consecutive limit ({max_consecutive}) on day {day_index + 1}."
                )
                break
            previous_period = period_index

    for (section_key, session_id), occurrences in lab_session_occurrences.items():
        if len(occurrences) != 2:
            issues.append(
                f"Lab session {session_id} in {section_key} does not occupy exactly 2 contiguous periods."
            )
            continue

        occurrences_sorted = sorted(occurrences, key=lambda item: (item[0], item[1]))
        first_day, first_period, first_cell = occurrences_sorted[0]
        second_day, second_period, second_cell = occurrences_sorted[1]

        if first_day != second_day or second_period != first_period + 1:
            issues.append(
                f"Lab session {session_id} in {section_key} is not in contiguous periods."
            )

        if int(first_cell.get("session_length", 0)) != 2 or int(second_cell.get("session_length", 0)) != 2:
            issues.append(
                f"Lab session {session_id} in {section_key} is not marked as double-period."
            )

        if {int(first_cell.get("session_part", 0)), int(second_cell.get("session_part", 0))} != {1, 2}:
            issues.append(
                f"Lab session {session_id} in {section_key} has invalid session parts."
            )

        key_fields = ("subject", "teacher", "room")
        for key_field in key_fields:
            if str(first_cell.get(key_field, "")) != str(second_cell.get(key_field, "")):
                issues.append(
                    f"Lab session {session_id} in {section_key} changes {key_field} across contiguous slots."
                )

    for bundle_key, occurrences in lab_bundle_occurrences.items():
        if len(occurrences) != 6:
            issues.append(
                f"Lab bundle {bundle_key} does not occupy exactly 3 grouped sections in 2 periods."
            )
            continue

        section_occurrences: Dict[str, List[Tuple[int, int, TimetableCell]]] = {}
        for section_key, day_index, period_index, cell in occurrences:
            section_occurrences.setdefault(section_key, []).append((day_index, period_index, cell))

        if len(section_occurrences) != 3:
            issues.append(
                f"Lab bundle {bundle_key} does not span exactly 3 section groups."
            )
            continue

        bundle_day: int | None = None
        bundle_start_period: int | None = None
        teacher_set: Set[str] = set()
        room_set: Set[str] = set()
        group_number_set: Set[int] = set()

        for section_key, section_items in section_occurrences.items():
            if len(section_items) != 2:
                issues.append(
                    f"Lab bundle {bundle_key} has invalid period count for {section_key}."
                )
                continue

            section_items = sorted(section_items, key=lambda item: (item[0], item[1]))
            first_day, first_period, first_cell = section_items[0]
            second_day, second_period, second_cell = section_items[1]

            if first_day != second_day or second_period != first_period + 1:
                issues.append(
                    f"Lab bundle {bundle_key} is not contiguous for {section_key}."
                )
                continue

            if not _cells_form_single_lab_block(first_cell, second_cell):
                issues.append(
                    f"Lab bundle {bundle_key} has invalid lab block metadata for {section_key}."
                )

            if bundle_day is None:
                bundle_day = first_day
                bundle_start_period = first_period
            elif first_day != bundle_day or first_period != bundle_start_period:
                issues.append(
                    f"Lab bundle {bundle_key} is not synchronized in the same time slot across groups."
                )

            teacher_set.add(str(first_cell.get("teacher", "")).strip())
            room_set.add(str(first_cell.get("room", "")).strip())

            _, group_number = _extract_section_group(str(first_cell.get("section", "")))
            if group_number is None:
                issues.append(
                    f"Lab bundle {bundle_key} section {section_key} is missing a G1/G2/G3 label."
                )
            else:
                group_number_set.add(group_number)

        if len(teacher_set) != 3:
            issues.append(
                f"Lab bundle {bundle_key} must use 3 distinct faculty members."
            )

        if len(room_set) != 3:
            issues.append(
                f"Lab bundle {bundle_key} must use 3 distinct lab rooms."
            )

        if group_number_set != {1, 2, 3}:
            issues.append(
                f"Lab bundle {bundle_key} must map to section groups G1, G2, and G3."
            )

    for index, fixed_slot in enumerate(normalized_fixed_slots, start=1):
        section_key = str(fixed_slot.get("section_key", ""))
        day_index = int(fixed_slot.get("day_index", -1))
        start_period = int(fixed_slot.get("start_period", -1))
        duration = int(fixed_slot.get("duration", 1) or 1)
        subject_name = str(fixed_slot.get("subject", ""))
        teacher_name = str(fixed_slot.get("teacher", ""))
        room_name = str(fixed_slot.get("room", "")).strip()

        if section_key not in table_by_section:
            issues.append(f"Fixed-slot constraint #{index} references unknown section {section_key}.")
            continue

        if day_index < 0 or day_index >= day_count or start_period < 0 or start_period + duration > period_count:
            issues.append(f"Fixed-slot constraint #{index} is out of timetable range.")
            continue

        table = table_by_section[section_key]
        matrix: TimetableMatrix = table["timetable"]
        for offset in range(duration):
            period_index = start_period + offset
            period_number = period_index + 1
            cell = matrix[day_index][period_index]

            if cell is None:
                issues.append(
                    f"Fixed-slot constraint #{index} is not satisfied: empty slot at day {day_index + 1}, period {period_number}."
                )
                continue

            if str(cell.get("subject", "")) != subject_name:
                issues.append(
                    f"Fixed-slot constraint #{index} subject mismatch at day {day_index + 1}, period {period_number}."
                )

            if str(cell.get("teacher", "")) != teacher_name:
                issues.append(
                    f"Fixed-slot constraint #{index} faculty mismatch at day {day_index + 1}, period {period_number}."
                )

            if room_name and str(cell.get("room", "")) != room_name:
                issues.append(
                    f"Fixed-slot constraint #{index} room mismatch at day {day_index + 1}, period {period_number}."
                )

    return issues


def export_multi_timetable_to_excel(
    days: List[str],
    periods: List[str],
    section_tables: List[SectionTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    class_tables = build_class_timetable_tables(section_tables)
    faculty_tables = build_faculty_timetable_tables(
        days,
        periods,
        section_tables,
        lunch_break_periods=lunch_break_periods,
    )
    room_tables = build_room_timetable_tables(
        days,
        periods,
        section_tables,
        lunch_break_periods=lunch_break_periods,
    )

    class_blocks: List[Tuple[str, TimetableMatrix]] = []
    for class_table in class_tables:
        for section_table in class_table["sections"]:
            class_blocks.append(
                (
                    build_section_key(
                        str(class_table["class_name"]),
                        str(class_table["standard"]),
                        str(section_table["section"]),
                    ),
                    section_table["timetable"],
                )
            )

    faculty_blocks: List[Tuple[str, TimetableMatrix]] = [
        (str(table["teacher"]), table["timetable"])
        for table in faculty_tables
    ]

    room_blocks: List[Tuple[str, TimetableMatrix]] = [
        (str(table["room"]), table["timetable"])
        for table in room_tables
    ]

    lab_blocks: List[Tuple[str, TimetableMatrix]] = []
    for table in room_tables:
        lab_matrix, has_lab = _extract_lab_only_matrix(table["timetable"])
        if has_lab:
            lab_blocks.append((str(table["room"]), lab_matrix))

    if not lab_blocks:
        for table in section_tables:
            lab_matrix, has_lab = _extract_lab_only_matrix(table["timetable"])
            if has_lab:
                lab_blocks.append((str(table["section_key"]), lab_matrix))

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _write_styled_timetable_sheet(
        workbook,
        sheet_name="CLASS",
        sheet_title="CLASS TIMETABLE",
        blocks=class_blocks,
        days=days,
        periods=periods,
        lunch_break_periods_by_day=lunch_by_day,
        view_mode="class",
    )
    _write_styled_timetable_sheet(
        workbook,
        sheet_name="FACULTY",
        sheet_title="FACULTY TIMETABLE",
        blocks=faculty_blocks,
        days=days,
        periods=periods,
        lunch_break_periods_by_day=lunch_by_day,
        view_mode="faculty",
    )
    _write_styled_timetable_sheet(
        workbook,
        sheet_name="LAB",
        sheet_title="LAB TIMETABLE",
        blocks=lab_blocks,
        days=days,
        periods=periods,
        lunch_break_periods_by_day=lunch_by_day,
        view_mode="lab",
    )
    _write_styled_timetable_sheet(
        workbook,
        sheet_name="ROOM",
        sheet_title="ROOM TIMETABLE",
        blocks=room_blocks,
        days=days,
        periods=periods,
        lunch_break_periods_by_day=lunch_by_day,
        view_mode="room",
    )

    workbook.save(output_path)


def _format_entity_label(class_name: str, standard: str, section: str) -> str:
    chunks: List[str] = []
    if class_name:
        chunks.append(class_name)
    if standard:
        chunks.append(f"Std {standard}")
    if section:
        chunks.append(f"Sec {section}")
    return " | ".join(chunks)


def _format_timetable_entry(cell: TimetableCell, view_mode: str) -> str:
    subject_label = str(cell.get("subject", "")).strip()
    if bool(cell.get("is_lab", False)) and subject_label:
        subject_label = f"{subject_label} [LAB]"

    teacher_name = str(cell.get("teacher", "")).strip()
    room_name = str(cell.get("room", "")).strip()
    section_label = _format_entity_label(
        str(cell.get("class_name", "")).strip(),
        str(cell.get("standard", "")).strip(),
        str(cell.get("section", "")).strip(),
    )

    if view_mode == "class":
        lines = [subject_label, teacher_name, room_name]
    elif view_mode == "faculty":
        lines = [subject_label, section_label, room_name]
    else:
        lines = [subject_label, teacher_name, section_label]

    return "\n".join(line for line in lines if line)


def _format_timetable_cell(cell: TimetableCell | None, view_mode: str) -> str:
    if cell is None:
        return ""

    if isinstance(cell, dict) and "conflicts" in cell:
        conflicts = [
            entry
            for entry in list(cell.get("conflicts", []))
            if isinstance(entry, dict)
        ]

        if view_mode == "lab":
            conflicts = [entry for entry in conflicts if bool(entry.get("is_lab", False))]

        if not conflicts:
            return ""

        lines = ["CONFLICT"]
        preview = conflicts[:2]
        preview_mode = "room" if view_mode == "lab" else view_mode
        for entry in preview:
            lines.append(_format_timetable_entry(entry, preview_mode))

        remaining = len(conflicts) - len(preview)
        if remaining > 0:
            lines.append(f"+{remaining} more")
        return "\n---\n".join(lines)

    if not isinstance(cell, dict):
        return str(cell)

    if view_mode == "lab" and not bool(cell.get("is_lab", False)):
        return ""

    effective_mode = "room" if view_mode == "lab" else view_mode
    return _format_timetable_entry(cell, effective_mode)


def _extract_lab_only_matrix(matrix: TimetableMatrix) -> Tuple[TimetableMatrix, bool]:
    filtered_matrix: TimetableMatrix = []
    has_lab_entries = False

    for row in matrix:
        filtered_row: List[TimetableCell | None] = []
        for cell in row:
            if cell is None:
                filtered_row.append(None)
                continue

            if isinstance(cell, dict) and "conflicts" in cell:
                lab_conflicts = [
                    entry
                    for entry in list(cell.get("conflicts", []))
                    if isinstance(entry, dict) and bool(entry.get("is_lab", False))
                ]
                if lab_conflicts:
                    filtered_row.append({"conflicts": lab_conflicts})
                    has_lab_entries = True
                else:
                    filtered_row.append(None)
                continue

            if isinstance(cell, dict) and bool(cell.get("is_lab", False)):
                filtered_row.append(cell)
                has_lab_entries = True
            else:
                filtered_row.append(None)

        filtered_matrix.append(filtered_row)

    return filtered_matrix, has_lab_entries


def _write_styled_timetable_sheet(
    workbook: Workbook,
    sheet_name: str,
    sheet_title: str,
    blocks: List[Tuple[str, TimetableMatrix]],
    days: List[str],
    periods: List[str],
    lunch_break_periods_by_day: Dict[int, Set[int]],
    view_mode: str,
) -> None:
    sheet = workbook.create_sheet(title=sheet_name)
    total_columns = len(periods) + 1

    thin_border = Border(
        left=Side(style="thin", color="9E9E9E"),
        right=Side(style="thin", color="9E9E9E"),
        top=Side(style="thin", color="9E9E9E"),
        bottom=Side(style="thin", color="9E9E9E"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    block_fill = PatternFill(fill_type="solid", fgColor="305496")
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    day_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    lunch_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    empty_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    data_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    conflict_fill = PatternFill(fill_type="solid", fgColor="F8CBAD")

    for column_index in range(1, total_columns + 1):
        column_letter = get_column_letter(column_index)
        if column_index == 1:
            sheet.column_dimensions[column_letter].width = 14
        else:
            sheet.column_dimensions[column_letter].width = 19

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = sheet.cell(row=1, column=1, value=sheet_title)
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = center_alignment
    title_cell.fill = title_fill
    title_cell.border = thin_border
    sheet.row_dimensions[1].height = 26

    current_row = 3

    if not blocks:
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_columns)
        empty_cell = sheet.cell(row=current_row, column=1, value="No timetable rows available.")
        empty_cell.font = Font(bold=True, color="5A5A5A")
        empty_cell.alignment = center_alignment
        empty_cell.fill = empty_fill
        empty_cell.border = thin_border
    else:
        for block_label, matrix in blocks:
            sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=total_columns,
            )
            block_cell = sheet.cell(row=current_row, column=1, value=block_label)
            block_cell.font = Font(color="FFFFFF", bold=True, size=11)
            block_cell.alignment = center_alignment
            block_cell.fill = block_fill
            block_cell.border = thin_border
            sheet.row_dimensions[current_row].height = 22
            current_row += 1

            header_cell = sheet.cell(row=current_row, column=1, value="DAY / PERIOD")
            header_cell.font = Font(bold=True)
            header_cell.alignment = center_alignment
            header_cell.fill = header_fill
            header_cell.border = thin_border

            for period_index, period_label in enumerate(periods, start=2):
                period_cell = sheet.cell(row=current_row, column=period_index, value=period_label)
                period_cell.font = Font(bold=True)
                period_cell.alignment = center_alignment
                period_cell.fill = header_fill
                period_cell.border = thin_border

            sheet.row_dimensions[current_row].height = 20
            current_row += 1

            for day_index, day_name in enumerate(days):
                day_cell = sheet.cell(row=current_row, column=1, value=day_name)
                day_cell.font = Font(bold=True)
                day_cell.alignment = center_alignment
                day_cell.fill = day_fill
                day_cell.border = thin_border

                for period_index in range(len(periods)):
                    period_number = period_index + 1
                    matrix_cell: TimetableCell | None = None
                    if day_index < len(matrix) and period_index < len(matrix[day_index]):
                        matrix_cell = matrix[day_index][period_index]

                    excel_cell = sheet.cell(row=current_row, column=period_index + 2)
                    excel_cell.alignment = center_alignment
                    excel_cell.border = thin_border

                    if _is_lunch_period(lunch_break_periods_by_day, day_index, period_number):
                        excel_cell.value = "LUNCH BREAK"
                        excel_cell.fill = lunch_fill
                        excel_cell.font = Font(bold=True, color="7A3F00")
                    else:
                        display_value = _format_timetable_cell(matrix_cell, view_mode)
                        excel_cell.value = display_value
                        if display_value.startswith("CONFLICT"):
                            excel_cell.fill = conflict_fill
                            excel_cell.font = Font(bold=True, color="9C0006")
                        elif display_value:
                            excel_cell.fill = data_fill
                            excel_cell.font = Font(size=10)
                        else:
                            excel_cell.fill = empty_fill
                            excel_cell.font = Font(color="7A7A7A", italic=True)

                sheet.row_dimensions[current_row].height = 46
                current_row += 1

            current_row += 2

    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0



def export_faculty_timetable_to_excel(
    days: List[str],
    periods: List[str],
    faculty_tables: List[FacultyTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for index, table in enumerate(faculty_tables):
            rows: List[List[str]] = []
            matrix: TimetableMatrix = table["timetable"]

            for day_index, row in enumerate(matrix):
                row_items: List[str] = []
                for period_index, cell in enumerate(row):
                    period_number = period_index + 1
                    if _is_lunch_period(lunch_by_day, day_index, period_number):
                        row_items.append("LUNCH BREAK")
                    elif cell is None:
                        row_items.append("")
                    elif isinstance(cell, dict) and "conflicts" in cell:
                        row_items.append("CONFLICT")
                    else:
                        subject_label = str(cell.get("subject", ""))
                        if bool(cell.get("is_lab", False)):
                            subject_label = f"{subject_label} [LAB]"

                        section_label = (
                            f"{cell.get('class_name', '')} | Std {cell.get('standard', '')} | Sec {cell.get('section', '')}"
                        )
                        row_items.append(
                            f"{subject_label}\n{section_label}\n{cell.get('room', '')}"
                        )
                rows.append(row_items)

            sheet_name = _sanitize_sheet_name(str(table["teacher"]), index, prefix="F")
            dataframe = pd.DataFrame(rows, index=days, columns=periods)
            dataframe.to_excel(writer, sheet_name=sheet_name)


def export_class_timetable_to_excel(
    days: List[str],
    periods: List[str],
    class_tables: List[ClassTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for index, class_table in enumerate(class_tables):
            rows: List[List[str]] = []

            for section_table in class_table["sections"]:
                section_name = str(section_table["section"])
                matrix: TimetableMatrix = section_table["timetable"]

                for day_index, day_name in enumerate(days):
                    row_items: List[str] = [section_name, day_name]
                    for period_index in range(len(periods)):
                        period_number = period_index + 1
                        cell = matrix[day_index][period_index]

                        if _is_lunch_period(lunch_by_day, day_index, period_number):
                            row_items.append("LUNCH BREAK")
                        elif cell is None:
                            row_items.append("")
                        else:
                            subject_label = str(cell.get("subject", ""))
                            if bool(cell.get("is_lab", False)):
                                subject_label = f"{subject_label} [LAB]"

                            row_items.append(
                                f"{subject_label}\n{cell.get('teacher', '')}\n{cell.get('room', '')}"
                            )

                    rows.append(row_items)

            columns = ["Section", "Day", *periods]
            sheet_name = _sanitize_sheet_name(str(class_table["class_key"]), index, prefix="C")
            dataframe = pd.DataFrame(rows, columns=columns)
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)


def export_room_timetable_to_excel(
    days: List[str],
    periods: List[str],
    room_tables: List[RoomTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for index, table in enumerate(room_tables):
            rows: List[List[str]] = []
            matrix: TimetableMatrix = table["timetable"]

            for day_index, row in enumerate(matrix):
                row_items: List[str] = []
                for period_index, cell in enumerate(row):
                    period_number = period_index + 1
                    if _is_lunch_period(lunch_by_day, day_index, period_number):
                        row_items.append("LUNCH BREAK")
                    elif cell is None:
                        row_items.append("")
                    elif isinstance(cell, dict) and "conflicts" in cell:
                        row_items.append("CONFLICT")
                    else:
                        subject_label = str(cell.get("subject", ""))
                        if bool(cell.get("is_lab", False)):
                            subject_label = f"{subject_label} [LAB]"

                        section_label = (
                            f"{cell.get('class_name', '')} | Std {cell.get('standard', '')} | Sec {cell.get('section', '')}"
                        )
                        row_items.append(
                            f"{subject_label}\n{cell.get('teacher', '')}\n{section_label}"
                        )
                rows.append(row_items)

            sheet_name = _sanitize_sheet_name(str(table["room"]), index, prefix="R")
            dataframe = pd.DataFrame(rows, index=days, columns=periods)
            dataframe.to_excel(writer, sheet_name=sheet_name)


def export_multi_timetable_to_csv(
    days: List[str],
    periods: List[str],
    section_tables: List[SectionTable],
    output_path: Path,
    lunch_break_periods: Set[int] | List[int] | Tuple[int, ...] | None = None,
) -> None:
    dataframe = multi_timetable_to_long_dataframe(
        days,
        periods,
        section_tables,
        lunch_break_periods=lunch_break_periods,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(output_path, index=False)


def export_multi_timetable_to_json(
    days: List[str],
    periods: List[str],
    section_tables: List[SectionTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
    summary: Dict[str, object] | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    lunch_set = _normalize_lunch_break_periods(
        lunch_break_periods,
        len(periods),
        len(days),
    )
    sections = [
        {
            "section_key": table["section_key"],
            "class_name": table["class_name"],
            "standard": table["standard"],
            "section": table["section"],
        }
        for table in section_tables
    ]

    payload: Dict[str, object] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "days": days,
        "periods": periods,
        "lunch_break_periods": sorted(lunch_set),
        "lunch_break_periods_by_day": {
            str(day_index + 1): sorted(lunch_by_day.get(day_index, set()))
            for day_index in range(len(days))
        },
        "sections": sections,
        "entries": multi_timetable_to_long_dataframe(
            days,
            periods,
            section_tables,
            lunch_break_periods=lunch_by_day,
        ).to_dict("records"),
    }
    if summary is not None:
        payload["summary"] = summary

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as output_file:
        json.dump(payload, output_file, indent=2)


def export_faculty_timetable_to_csv(
    days: List[str],
    periods: List[str],
    faculty_tables: List[FacultyTable],
    output_path: Path,
    lunch_break_periods: Set[int] | List[int] | Tuple[int, ...] | None = None,
) -> None:
    dataframe = faculty_timetable_to_long_dataframe(
        days,
        periods,
        faculty_tables,
        lunch_break_periods=lunch_break_periods,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(output_path, index=False)


def export_class_timetable_to_csv(
    days: List[str],
    periods: List[str],
    class_tables: List[ClassTable],
    output_path: Path,
    lunch_break_periods: Set[int] | List[int] | Tuple[int, ...] | None = None,
) -> None:
    dataframe = class_timetable_to_long_dataframe(
        days,
        periods,
        class_tables,
        lunch_break_periods=lunch_break_periods,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(output_path, index=False)


def export_room_timetable_to_csv(
    days: List[str],
    periods: List[str],
    room_tables: List[RoomTable],
    output_path: Path,
    lunch_break_periods: Set[int] | List[int] | Tuple[int, ...] | None = None,
) -> None:
    dataframe = room_timetable_to_long_dataframe(
        days,
        periods,
        room_tables,
        lunch_break_periods=lunch_break_periods,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(output_path, index=False)


def export_faculty_timetable_to_json(
    days: List[str],
    periods: List[str],
    faculty_tables: List[FacultyTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
    summary: Dict[str, object] | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    lunch_set = _normalize_lunch_break_periods(
        lunch_break_periods,
        len(periods),
        len(days),
    )
    payload: Dict[str, object] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "days": days,
        "periods": periods,
        "lunch_break_periods": sorted(lunch_set),
        "lunch_break_periods_by_day": {
            str(day_index + 1): sorted(lunch_by_day.get(day_index, set()))
            for day_index in range(len(days))
        },
        "faculty": [
            {"teacher": str(table["teacher"])}
            for table in faculty_tables
        ],
        "entries": faculty_timetable_to_long_dataframe(
            days,
            periods,
            faculty_tables,
            lunch_break_periods=lunch_by_day,
        ).to_dict("records"),
    }

    if summary is not None:
        payload["summary"] = {
            "faculty_count": summary.get("faculty_count", 0),
            "faculty_stats": summary.get("faculty_stats", {}),
            "teacher_loads": summary.get("teacher_loads", {}),
        }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as output_file:
        json.dump(payload, output_file, indent=2)


def export_class_timetable_to_json(
    days: List[str],
    periods: List[str],
    class_tables: List[ClassTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
    summary: Dict[str, object] | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    lunch_set = _normalize_lunch_break_periods(
        lunch_break_periods,
        len(periods),
        len(days),
    )
    payload: Dict[str, object] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "days": days,
        "periods": periods,
        "lunch_break_periods": sorted(lunch_set),
        "lunch_break_periods_by_day": {
            str(day_index + 1): sorted(lunch_by_day.get(day_index, set()))
            for day_index in range(len(days))
        },
        "classes": [
            {
                "class_key": str(class_table["class_key"]),
                "class_name": str(class_table["class_name"]),
                "standard": str(class_table["standard"]),
                "sections": [
                    {
                        "section": str(section_table["section"]),
                        "section_key": str(section_table["section_key"]),
                    }
                    for section_table in class_table["sections"]
                ],
            }
            for class_table in class_tables
        ],
        "entries": class_timetable_to_long_dataframe(
            days,
            periods,
            class_tables,
            lunch_break_periods=lunch_by_day,
        ).to_dict("records"),
    }

    if summary is not None:
        payload["summary"] = {
            "class_count": summary.get("class_count", 0),
            "class_stats": summary.get("class_stats", {}),
            "section_count": summary.get("section_count", 0),
        }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as output_file:
        json.dump(payload, output_file, indent=2)


def export_room_timetable_to_json(
    days: List[str],
    periods: List[str],
    room_tables: List[RoomTable],
    output_path: Path,
    lunch_break_periods: Dict[int, Set[int] | List[int] | Tuple[int, ...] | int]
    | Set[int]
    | List[int]
    | Tuple[int, ...]
    | None = None,
    summary: Dict[str, object] | None = None,
) -> None:
    lunch_by_day = _normalize_lunch_break_periods_by_day(
        lunch_break_periods,
        len(days),
        len(periods),
    )
    lunch_set = _normalize_lunch_break_periods(
        lunch_break_periods,
        len(periods),
        len(days),
    )
    payload: Dict[str, object] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "days": days,
        "periods": periods,
        "lunch_break_periods": sorted(lunch_set),
        "lunch_break_periods_by_day": {
            str(day_index + 1): sorted(lunch_by_day.get(day_index, set()))
            for day_index in range(len(days))
        },
        "rooms": [
            {"room": str(table["room"])}
            for table in room_tables
        ],
        "entries": room_timetable_to_long_dataframe(
            days,
            periods,
            room_tables,
            lunch_break_periods=lunch_by_day,
        ).to_dict("records"),
    }

    if summary is not None:
        payload["summary"] = {
            "room_count": summary.get("room_count", 0),
            "room_stats": summary.get("room_stats", {}),
            "room_loads": summary.get("room_loads", {}),
        }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as output_file:
        json.dump(payload, output_file, indent=2)
